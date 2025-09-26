import os
from flask import Flask, render_template, jsonify, request, send_file, redirect, url_for, flash, session, make_response
import json
import pandas as pd
from app import app, db
import io
from auth_supabase import supabase_auth_manager as auth_manager, login_required, admin_required
from flask_migrate import Migrate
from models import User, Revenda, Vendedor
from datetime import datetime
from openpyxl.styles import PatternFill, Font
import openpyxl # Import openpyxl

# Initialize Migration
migrate = Migrate(app, db)



# Load crop data
def load_crop_data():
    try:
        with open('data/crop_data_static.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print("Arquivo crop_data_static.json não encontrado")
        return {}
    except Exception as e:
        print(f"Erro ao carregar dados: {e}")
        return {}

# Load fertilizer data
def load_fertilizer_data():
    try:
        with open('data/fertilizer_data_static_corrigido.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print("Arquivo fertilizer_data_static_corrigido.json não encontrado")
        return {}
    except Exception as e:
        print(f"Erro ao carregar dados de fertilizantes: {e}")
        return {}

# Load static data files
try:
    with open('data/crop_data_static.json', 'r', encoding='utf-8') as f:
        CROP_DATA = json.load(f)
    print(f"Loaded crop data with {len(CROP_DATA)} crops")
except Exception as e:
    print(f"Error loading crop data: {e}")
    CROP_DATA = {}

try:
    with open('data/fertilizer_data_static_corrigido.json', 'r', encoding='utf-8') as f:
        FERTILIZER_DATA = json.load(f)
    print(f"Loaded fertilizer data with {len(FERTILIZER_DATA)} categories")
except Exception as e:
    print(f"Error loading fertilizer data: {e}")
    FERTILIZER_DATA = {}

try:
    with open('data/agrotoxico_data_static.json', 'r', encoding='utf-8') as f:
        AGROTOXICO_DATA = json.load(f)
    print(f"Loaded agrotoxico data with {len(AGROTOXICO_DATA)} categories")
except Exception as e:
    print(f"Error loading agrotoxico data: {e}")
    AGROTOXICO_DATA = {}

try:
    with open('data/consultoria_tecnica_data_static.json', 'r', encoding='utf-8') as f:
        CONSULTORIA_DATA = json.load(f)
    print(f"Loaded consultoria tecnica data with {len(CONSULTORIA_DATA)} categories")
except Exception as e:
    print(f"Error loading consultoria tecnica data: {e}")
    CONSULTORIA_DATA = {}

try:
    with open('data/corretivos_data_static.json', 'r', encoding='utf-8') as f:
        CORRETIVOS_DATA = json.load(f)
    print(f"Loaded corretivos data with {len(CORRETIVOS_DATA)} categories")
except Exception as e:
    print(f"Error loading corretivos data: {e}")
    CORRETIVOS_DATA = {}

try:
    with open('data/despesa_data_static.json', 'r', encoding='utf-8') as f:
        DESPESA_DATA = json.load(f)
    print(f"Loaded despesa data with {len(DESPESA_DATA)} categories")
except Exception as e:
    print(f"Error loading despesa data: {e}")
    DESPESA_DATA = {}

try:
    with open('data/escolaridade_data_static.json', 'r', encoding='utf-8') as f:
        ESCOLARIDADE_DATA = json.load(f)
    print(f"Loaded escolaridade data with {len(ESCOLARIDADE_DATA)} categories")
except Exception as e:
    print(f"Error loading escolaridade data: {e}")
    ESCOLARIDADE_DATA = {}

try:
    with open('data/receita_data_static.json', 'r', encoding='utf-8') as f:
        RECEITA_DATA = json.load(f)
    print(f"Loaded receita data with {len(RECEITA_DATA)} categories")
except Exception as e:
    print(f"Error loading receita data: {e}")
    RECEITA_DATA = {}

@app.route('/')
@login_required
def index():
    user = auth_manager.get_current_user()
    return render_template('index.html', user=user)

@app.route('/analysis')
@login_required
def analysis():
    user = auth_manager.get_current_user()
    return render_template('analysis.html', user=user)

@app.route('/api/brazilian-states')
def get_states():
    try:
        # Brazilian states
        states = [
            {'code': 'AC', 'name': 'Acre'},
            {'code': 'AL', 'name': 'Alagoas'},
            {'code': 'AP', 'name': 'Amapá'},
            {'code': 'AM', 'name': 'Amazonas'},
            {'code': 'BA', 'name': 'Bahia'},
            {'code': 'CE', 'name': 'Ceará'},
            {'code': 'DF', 'name': 'Distrito Federal'},
            {'code': 'ES', 'name': 'Espírito Santo'},
            {'code': 'GO', 'name': 'Goiás'},
            {'code': 'MA', 'name': 'Maranhão'},
            {'code': 'MT', 'name': 'Mato Grosso'},
            {'code': 'MS', 'name': 'Mato Grosso do Sul'},
            {'code': 'MG', 'name': 'Minas Gerais'},
            {'code': 'PA', 'name': 'Pará'},
            {'code': 'PB', 'name': 'Paraíba'},
            {'code': 'PR', 'name': 'Paraná'},
            {'code': 'PE', 'name': 'Pernambuco'},
            {'code': 'PI', 'name': 'Piauí'},
            {'code': 'RJ', 'name': 'Rio de Janeiro'},
            {'code': 'RN', 'name': 'Rio Grande do Norte'},
            {'code': 'RS', 'name': 'Rio Grande do Sul'},
            {'code': 'RO', 'name': 'Rondônia'},
            {'code': 'RR', 'name': 'Roraima'},
            {'code': 'SC', 'name': 'Santa Catarina'},
            {'code': 'SP', 'name': 'São Paulo'},
            {'code': 'SE', 'name': 'Sergipe'},
            {'code': 'TO', 'name': 'Tocantins'}
        ]

        return jsonify({
            'success': True,
            'states': states
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/statistics')
def get_statistics():
    try:
        # CROP_DATA structure: {crop_name: {municipality_code: {data}}}
        total_crops = len(CROP_DATA)
        total_fertilizer_categories = len(FERTILIZER_DATA)
        total_agrotoxico_categories = len(AGROTOXICO_DATA)
        total_consultoria_categories = len(CONSULTORIA_DATA)
        total_corretivos_categories = len(CORRETIVOS_DATA)
        total_despesa_categories = len(DESPESA_DATA)
        total_escolaridade_categories = len(ESCOLARIDADE_DATA)
        total_receita_categories = len(RECEITA_DATA)

        # Count unique municipalities across all crops
        all_municipalities = set()
        for crop_data in CROP_DATA.values():
            all_municipalities.update(crop_data.keys())

        # Count unique municipalities across all fertilizer categories
        all_fertilizer_municipalities = set()
        for fertilizer_data in FERTILIZER_DATA.values():
            all_fertilizer_municipalities.update(fertilizer_data.keys())

        total_municipalities = len(all_municipalities)
        total_fertilizer_municipalities = len(all_fertilizer_municipalities)

        # Calculate total establishments for fertilizer data
        total_establishments = 0
        if 'Total Estabelecimentos' in FERTILIZER_DATA:
            for municipality_data in FERTILIZER_DATA['Total Estabelecimentos'].values():
                if isinstance(municipality_data, dict) and 'value' in municipality_data:
                    total_establishments += municipality_data.get('value', 0)

        return jsonify({
            'success': True,
            'statistics': {
                'total_crops': total_crops,
                'total_municipalities': total_municipalities,
                'total_fertilizer_categories': total_fertilizer_categories,
                'total_agrotoxico_categories': total_agrotoxico_categories,
                'total_consultoria_categories': total_consultoria_categories,
                'total_corretivos_categories': total_corretivos_categories,
                'total_despesa_categories': total_despesa_categories,
                'total_escolaridade_categories': total_escolaridade_categories,
                'total_receita_categories': total_receita_categories,
                'total_establishments': total_establishments,
                'data_sources': {
                    'culturas': total_crops,
                    'fertilizantes': total_fertilizer_categories,
                    'agrotoxicos': total_agrotoxico_categories,
                    'consultoria_tecnica': total_consultoria_categories,
                    'corretivos': total_corretivos_categories,
                    'despesas': total_despesa_categories,
                    'escolaridade': total_escolaridade_categories,
                    'receitas': total_receita_categories
                }
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/crops')
def get_crops():
    try:
        # CROP_DATA structure: {crop_name: {municipality_code: {data}}}
        sorted_crops = sorted(list(CROP_DATA.keys()))
        return jsonify({
            'success': True,
            'crops': sorted_crops
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/fertilizer-categories')
def get_fertilizer_categories():
    try:
        # FERTILIZER_DATA structure: {category_name: {municipality_code: {data}}}
        categories = sorted(list(FERTILIZER_DATA.keys()))
        return jsonify({
            'success': True,
            'categories': categories
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/agrotoxico/categories')
def get_agrotoxico_categories():
    try:
        categories = list(AGROTOXICO_DATA.keys())
        return jsonify({
            "success": True,
            "categories": categories,
            "total": len(categories)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/consultoria/categories')
def get_consultoria_categories():
    try:
        categories = list(CONSULTORIA_DATA.keys())
        return jsonify({
            "success": True,
            "categories": categories,
            "total": len(categories)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/corretivos/categories')
def get_corretivos_categories():
    try:
        categories = list(CORRETIVOS_DATA.keys())
        return jsonify({
            "success": True,
            "categories": categories,
            "total": len(categories)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/despesa/categories')
def get_despesa_categories():
    try:
        categories = list(DESPESA_DATA.keys())
        return jsonify({
            "success": True,
            "categories": categories,
            "total": len(categories)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/escolaridade/categories')
def get_escolaridade_categories():
    try:
        categories = list(ESCOLARIDADE_DATA.keys())
        return jsonify({
            "success": True,
            "categories": categories,
            "total": len(categories)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/receita/categories')
def get_receita_categories():
    try:
        categories = list(RECEITA_DATA.keys())
        return jsonify({
            "success": True,
            "categories": categories,
            "total": len(categories)
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/fertilizer-data/<category_name>')
def get_fertilizer_data(category_name):
    try:
        # Busca exata primeiro
        if category_name in FERTILIZER_DATA:
            all_fertilizer_data = FERTILIZER_DATA[category_name]

            # Filtrar apenas municípios válidos (códigos IBGE reais de municípios)
            fertilizer_municipalities = {}
            for municipality_code, municipality_data in all_fertilizer_data.items():
                municipality_code_str = str(municipality_code)
                municipality_name = municipality_data.get('municipality_name', '').lower()

                # Verificar se é um código de município válido
                if (len(municipality_code_str) == 7 and 
                    municipality_code_str.isdigit() and
                    municipality_code_str[0] in '12345' and  # Códigos reais começam com 1-5
                    municipality_data.get('municipality_name') and
                    # Excluir nomes que indicam regiões/agregações
                    not any(keyword in municipality_name for keyword in [
                        'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                        'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                        'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                        'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                    ])):
                    # Padronizar o nome do campo para compatibilidade
                    fertilizer_data = {
                        'municipality_name': municipality_data.get('municipality_name'),
                        'state_code': municipality_data.get('state_code'),
                        'harvested_area': municipality_data.get('value', 0),  # Mapear 'value' para 'harvested_area'
                        'unit': municipality_data.get('unit', 'un')
                    }
                    fertilizer_municipalities[municipality_code] = fertilizer_data

            return jsonify({
                'success': True,
                'data': fertilizer_municipalities,
                'data_type': 'fertilizer'
            })

        return jsonify({'success': False, 'error': 'Categoria de fertilizantes não encontrada'})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/fertilizer/<category>')
def get_fertilizer_data_alt(category):
    """Endpoint alternativo para fertilizantes"""
    return get_fertilizer_data(category)

@app.route('/api/agrotoxico/<category>')
def get_agrotoxico_data(category):
    try:
        if category in AGROTOXICO_DATA:
            return jsonify({
                'success': True,
                'data': AGROTOXICO_DATA[category],
                'type': 'agrotoxico'
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Categoria de agrotóxico "{category}" não encontrada'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/consultoria/<category>')
def get_consultoria_data(category):
    try:
        if category in CONSULTORIA_DATA:
            return jsonify({
                'success': True,
                'data': CONSULTORIA_DATA[category],
                'category': category
            })
        else:
            return jsonify({
                'success': False,
                'error': f"Categoria '{category}' não encontrada"
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/corretivos/<category>')
def get_corretivos_data(category):
    try:
        if category in CORRETIVOS_DATA:
            return jsonify({
                'success': True,
                'data': CORRETIVOS_DATA[category],
                'category': category
            })
        else:
            return jsonify({
                'success': False,
                'error': f"Categoria '{category}' não encontrada"
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/despesa/<category>')
def get_despesa_data(category):
    try:
        if category in DESPESA_DATA:
            return jsonify({
                'success': True,
                'data': DESPESA_DATA[category],
                'type': 'despesa'
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Categoria de despesa "{category}" não encontrada'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/escolaridade/<category>')
def get_escolaridade_data(category):
    try:
        if category in ESCOLARIDADE_DATA:
            return jsonify({
                'success': True,
                'data': ESCOLARIDADE_DATA[category],
                'category': category
            })
        else:
            return jsonify({
                'success': False,
                'error': f"Categoria '{category}' não encontrada"
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/receita/<category>')
def get_receita_data(category):
    try:
        if category in RECEITA_DATA:
            return jsonify({
                'success': True,
                'data': RECEITA_DATA[category],
                'type': 'receita'
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Categoria de receita "{category}" não encontrada'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/crop-data/<crop_name>')
def get_crop_data(crop_name):
    try:
        # Busca exata primeiro
        if crop_name in CROP_DATA:
            all_crop_data = CROP_DATA[crop_name]

            # Filtrar apenas municípios válidos (códigos IBGE reais de municípios)
            crop_municipalities = {}
            for municipality_code, municipality_data in all_crop_data.items():
                municipality_code_str = str(municipality_code)
                municipality_name = municipality_data.get('municipality_name', '').lower()

                # Verificar se é um código de município válido
                # Códigos de município IBGE começam com 1-5 e têm 7 dígitos
                # Excluir códigos que começam com 0 (são agregações regionais)
                if (len(municipality_code_str) == 7 and 
                    municipality_code_str.isdigit() and
                    municipality_code_str[0] in '12345' and  # Códigos reais começam com 1-5
                    municipality_data.get('municipality_name') and
                    # Excluir nomes que indicam regiões/agregações
                    not any(keyword in municipality_name for keyword in [
                        'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                        'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                        'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                        'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                    ]) and
                    # Excluir nomes muito genéricos ou que são claramente regiões
                    municipality_name not in [
                        'alto teles pires', 'sudeste mato-grossense', 'parecis', 'barreiras',
                        'dourados', 'norte mato-grossense', 'portal da amazônia'
                    ]):
                    crop_municipalities[municipality_code] = municipality_data

            # Debug: Encontrar o maior produtor para verificação
            if crop_municipalities:
                # Ordenar por área colhida para debug
                sorted_municipalities = sorted(crop_municipalities.items(), 
                                             key=lambda x: float(x[1].get('harvested_area', 0)), 
                                             reverse=True)
                max_municipality = sorted_municipalities[0]
                print(f"Debug - Maior produtor de {crop_name} (apenas municípios): {max_municipality[1].get('municipality_name')} ({max_municipality[1].get('state_code')}) - {max_municipality[1].get('harvested_area')} hectares")

                # Mostrar top 5 municípios para verificação
                print(f"Debug - Top 5 municípios produtores de {crop_name}:")
                for i, (code, data) in enumerate(sorted_municipalities[:5]):
                    print(f"  {i+1}. {data.get('municipality_name')} ({data.get('state_code')}): {data.get('harvested_area')} ha - Código: {code}")
            else:
                print(f"Debug - Nenhum município válido encontrado para {crop_name}")

            return jsonify({
                'success': True,
                'data': crop_municipalities
            })

        # Busca similar se não encontrar exata
        crop_name_lower = crop_name.lower()
        similar_crops = []

        for available_crop in CROP_DATA.keys():
            if crop_name_lower in available_crop.lower() or available_crop.lower() in crop_name_lower:
                similar_crops.append(available_crop)

        if similar_crops:
            # Usar a primeira cultura similar encontrada
            best_match = similar_crops[0]
            all_crop_data = CROP_DATA[best_match]

            # Filtrar apenas municípios válidos (códigos IBGE reais de municípios)
            crop_municipalities = {}
            for municipality_code, municipality_data in all_crop_data.items():
                municipality_code_str = str(municipality_code)
                municipality_name = municipality_data.get('municipality_name', '').lower()

                # Verificar se é um código de município válido
                # Códigos de município IBGE começam com 1-5 e têm 7 dígitos
                # Excluir códigos que começam com 0 (são agregações regionais)
                if (len(municipality_code_str) == 7 and 
                    municipality_code_str.isdigit() and
                    municipality_code_str[0] in '12345' and  # Códigos reais começam com 1-5
                    municipality_data.get('municipality_name') and
                    # Excluir nomes que indicam regiões/agregações
                    not any(keyword in municipality_name for keyword in [
                        'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                        'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                        'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                        'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                    ]) and
                    # Excluir nomes muito genéricos ou que são claramente regiões
                    municipality_name not in [
                        'alto teles pires', 'sudeste mato-grossense', 'parecis', 'barreiras',
                        'dourados', 'norte mato-grossense', 'portal da amazônia'
                    ]):
                    crop_municipalities[municipality_code] = municipality_data

            # Debug: Encontrar o maior produtor para verificação
            if crop_municipalities:
                # Ordenar por área colhida para debug
                sorted_municipalities = sorted(crop_municipalities.items(), 
                                             key=lambda x: float(x[1].get('harvested_area', 0)), 
                                             reverse=True)
                max_municipality = sorted_municipalities[0]
                print(f"Debug - Maior produtor de {best_match} (apenas municípios): {max_municipality[1].get('municipality_name')} ({max_municipality[1].get('state_code')}) - {max_municipality[1].get('harvested_area')} hectares")

                # Mostrar top 5 municípios para verificação
                print(f"Debug - Top 5 municípios produtores de {best_match}:")
                for i, (code, data) in enumerate(sorted_municipalities[:5]):
                    print(f"  {i+1}. {data.get('municipality_name')} ({data.get('state_code')}): {data.get('harvested_area')} ha - Código: {code}")
            else:
                print(f"Debug - Nenhum município válido encontrado para {best_match}")

            return jsonify({
                'success': True,
                'data': crop_municipalities,
                'matched_crop': best_match
            })

        return jsonify({'success': False, 'error': 'Cultura não encontrada'})

    except Exception as e:
        print(f"Erro crítico em get_crop_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Erro interno do servidor: {str(e)}'}), 500

@app.route('/api/crop-chart-data/<crop_name>')
def get_crop_chart_data(crop_name):
    try:
        if crop_name not in CROP_DATA:
            return jsonify({'success': False, 'error': 'Cultura não encontrada'})

        crop_municipalities = []
        for municipality_code, municipality_data in CROP_DATA[crop_name].items():
            # Filtrar apenas municípios válidos (códigos IBGE reais de municípios)
            municipality_code_str = str(municipality_code)
            municipality_name = municipality_data.get('municipality_name', '').lower()

            # Verificar se é um código de município válido
            # Códigos de município IBGE começam com 1-5 e têm 7 dígitos
            # Excluir códigos que começam com 0 (são agregações regionais)
            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and  # Códigos reais começam com 1-5
                municipality_data.get('municipality_name') and
                # Excluir nomes que indicam regiões/agregações
                not any(keyword in municipality_name for keyword in [
                    'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                    'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                    'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                    'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                ]) and
                # Excluir nomes muito genéricos ou que são claramente regiões
                municipality_name not in [
                    'alto teles pires', 'sudeste mato-grossense', 'parecis', 'barreiras',
                    'dourados', 'norte mato-grossense', 'portal da amazônia'
                ]):
                crop_municipalities.append({
                    'municipality_name': municipality_data.get('municipality_name', 'Desconhecido'),
                    'state_code': municipality_data.get('state_code', 'XX'),
                    'harvested_area': municipality_data.get('harvested_area', 0)
                })

        # Sort by harvested area and take top 20
        crop_municipalities.sort(key=lambda x: x['harvested_area'], reverse=True)
        top_20 = crop_municipalities[:20]

        chart_data = {
            'labels': [f"{muni['municipality_name']} ({muni['state_code']})" for muni in top_20],
            'data': [muni['harvested_area'] for muni in top_20]
        }

        return jsonify({
            'success': True,
            'chart_data': chart_data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/analysis/statistical-summary/<crop_name>')
def get_statistical_summary(crop_name):
    try:
        if crop_name not in CROP_DATA:
            return jsonify({'success': False, 'error': 'Cultura não encontrada'})

        # Filtrar apenas municípios válidos (códigos IBGE reais de municípios)
        values = []
        for municipality_code, data in CROP_DATA[crop_name].items():
            municipality_code_str = str(municipality_code)
            municipality_name = data.get('municipality_name', '').lower()

            # Verificar se é um código de município válido
            # Códigos de município IBGE começam com 1-5 e têm 7 dígitos
            # Excluir códigos que começam com 0 (são agregações regionais)
            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and  # Códigos reais começam com 1-5
                data.get('municipality_name') and
                # Excluir nomes que indicam regiões/agregações
                not any(keyword in municipality_name for keyword in [
                    'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                    'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                    'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                    'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                ]) and
                # Excluir nomes muito genéricos ou que são claramente regiões
                municipality_name not in [
                    'alto teles pires', 'sudeste mato-grossense', 'parecis', 'barreiras',
                    'dourados', 'norte mato-grossense', 'portal da amazônia'
                ]):
                values.append(data['harvested_area'])

        if not values:
            return jsonify({'success': False, 'error': 'Nenhum município válido encontrado para esta cultura'})

        import statistics
        summary = {
            'mean': statistics.mean(values),
            'median': statistics.median(values),
            'mode': statistics.mode(values) if len(set(values)) < len(values) else None,
            'std_dev': statistics.stdev(values) if len(values) > 1 else 0,
            'min': min(values),
            'max': max(values),
            'q1': statistics.quantiles(values, n=4)[0] if len(values) >= 4 else None,
            'q3': statistics.quantiles(values, n=4)[2] if len(values) >= 4 else None,
            'total': sum(values),
            'count': len(values)
        }

        return jsonify({
            'success': True,
            'summary': summary
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/analysis/by-state/<crop_name>')
def get_analysis_by_state(crop_name):
    try:
        if crop_name not in CROP_DATA:
            return jsonify({'success': False, 'error': 'Cultura não encontrada'})

        states_data = {}
        for municipality_code, municipality_data in CROP_DATA[crop_name].items():
            # Filtrar apenas municípios válidos (códigos IBGE reais de municípios)
            municipality_code_str = str(municipality_code)
            municipality_name = municipality_data.get('municipality_name', '').lower()

            # Verificar se é um código de município válido
            # Códigos de município IBGE começam com 1-5 e têm 7 dígitos
            # Excluir códigos que começam com 0 (são agregações regionais)
            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and  # Códigos reais começam com 1-5
                municipality_data.get('municipality_name') and
                # Excluir nomes que indicam regiões/agregações
                not any(keyword in municipality_name for keyword in [
                    'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                    'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                    'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                    'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                ]) and
                # Excluir nomes muito genéricos ou que são claramente regiões
                municipality_name not in [
                    'alto teles pires', 'sudeste mato-grossense', 'parecis', 'barreiras',
                    'dourados', 'norte mato-grossense', 'portal da amazônia'
                ]):

                state = municipality_data.get('state_code', 'XX')
                area = municipality_data.get('harvested_area', 0)

                if state not in states_data:
                    states_data[state] = {
                        'total_area': 0,
                        'municipalities_count': 0,
                        'max_area': 0,
                        'municipalities': []
                    }

                states_data[state]['total_area'] += area
                states_data[states_data[state]['municipalities_count']] += 1
                states_data[state]['max_area'] = max(states_data[state]['max_area'], area)
                states_data[state]['municipalities'].append({
                    'name': municipality_data.get('municipality_name'),
                    'area': area
                })

        # Calculate averages
        for state_data in states_data.values():
            state_data['average_area'] = state_data['total_area'] / state_data['municipalities_count']

        return jsonify({
            'success': True,
            'states_data': states_data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/analysis/comparison/<crop1>/<crop2>')
def get_crop_comparison(crop1, crop2):
    try:
        if crop1 not in CROP_DATA or crop2 not in CROP_DATA:
            return jsonify({'success': False, 'error': 'Uma ou ambas culturas não encontradas'})

        # Get common municipalities
        common_municipalities = set(CROP_DATA[crop1].keys()) & set(CROP_DATA[crop2].keys())

        comparison_data = []
        for muni_code in common_municipalities:
            data1 = CROP_DATA[crop1][muni_code]
            data2 = CROP_DATA[crop2][muni_code]

            comparison_data.append({
                'municipality_code': muni_code,
                'municipality_name': data1.get('municipality_name'),
                'state_code': data1.get('state_code'),
                'crop1_area': data1.get('harvested_area', 0),
                'crop2_area': data2.get('harvested_area', 0),
                'ratio': data1.get('harvested_area', 0) / max(data2.get('harvested_area', 1), 1)
            })

        return jsonify({
            'success': True,
            'crop1': crop1,
            'crop2': crop2,
            'comparison_data': comparison_data,
            'common_municipalities': len(common_municipalities)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/brazilian-states')
def get_brazilian_states():
    """Get list of Brazilian states"""
    try:
        states = [
            {'code': 'AC', 'name': 'Acre'},
            {'code': 'AL', 'name': 'Alagoas'},
            {'code': 'AP', 'name': 'Amapá'},
            {'code': 'AM', 'name': 'Amazonas'},
            {'code': 'BA', 'name': 'Bahia'},
            {'code': 'CE', 'name': 'Ceará'},
            {'code': 'DF', 'name': 'Distrito Federal'},
            {'code': 'ES', 'name': 'Espírito Santo'},
            {'code': 'GO', 'name': 'Goiás'},
            {'code': 'MA', 'name': 'Maranhão'},
            {'code': 'MT', 'name': 'Mato Grosso'},
            {'code': 'MS', 'name': 'Mato Grosso do Sul'},
            {'code': 'MG', 'name': 'Minas Gerais'},
            {'code': 'PA', 'name': 'Pará'},
            {'code': 'PB', 'name': 'Paraíba'},
            {'code': 'PR', 'name': 'Paraná'},
            {'code': 'PE', 'name': 'Pernambuco'},
            {'code': 'PI', 'name': 'Piauí'},
            {'code': 'RJ', 'name': 'Rio de Janeiro'},
            {'code': 'RN', 'name': 'Rio Grande do Norte'},
            {'code': 'RS', 'name': 'Rio Grande do Sul'},
            {'code': 'RO', 'name': 'Rondônia'},
            {'code': 'RR', 'name': 'Roraima'},
            {'code': 'SC', 'name': 'Santa Catarina'},
            {'code': 'SP', 'name': 'São Paulo'},
            {'code': 'SE', 'name': 'Sergipe'},
            {'code': 'TO', 'name': 'Tocantins'}
        ]

        return jsonify({
            'success': True,
            'states': states
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/export/crops')
def export_crops_data():
    """Export complete crop database as Excel file"""
    return export_complete_data()

@app.route('/api/export/fertilizers')
def export_fertilizers_data():
    """Export complete fertilizer database as Excel file"""
    return export_complete_fertilizer_data()

@app.route('/api/export/complete-data')
def export_complete_data():
    """Export complete crop data as Excel file"""
    try:
        # Load the original Excel file
        excel_path = os.path.join('data', 'ibge_2023_hectares_colhidos.xlsx')

        if not os.path.exists(excel_path):
            # Try alternative path
            excel_path = os.path.join('attached_assets', 'IBGE - 2023 - BRASIL HECTARES COLHIDOS_1752980032040.xlsx')

        if not os.path.exists(excel_path):
            return jsonify({'success': False, 'error': 'Arquivo de dados não encontrado'}), 404

        # Read the Excel file
        df = pd.read_excel(excel_path)

        # Create a BytesIO object to store the Excel file
        output = io.BytesIO()

        # Write to Excel
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Culturas IBGE 2023', index=False)

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='base_completa_culturas_ibge_2023.xlsx'
        )

    except Exception as e:
        print(f"Erro ao exportar dados: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/complete-fertilizer-data')
def export_complete_fertilizer_data():
    """Export complete fertilizer database as Excel file"""
    try:
        # Preparar todos os dados de fertilizantes para exportação
        all_fertilizer_data = []

        for category_name, category_data in FERTILIZER_DATA.items():
            for municipality_code, municipality_data in category_data.items():
                # Filtrar apenas municípios válidos
                municipality_code_str = str(municipality_code)
                municipality_name = municipality_data.get('municipality_name', '').lower()

                if (len(municipality_code_str) == 7 and 
                    municipality_code_str.isdigit() and
                    municipality_code_str[0] in '12345' and
                    municipality_data.get('municipality_name') and
                    not any(keyword in municipality_name for keyword in [
                        'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                        'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                        'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                        'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                    ])):

                    all_fertilizer_data.append({
                        'Código IBGE': municipality_code,
                        'Município': municipality_data.get('municipality_name', 'Desconhecido'),
                        'UF': municipality_data.get('state_code', 'XX'),
                        'Categoria': category_name,
                        'Valor': municipality_data.get('value', 0),
                        'Unidade': 'estabelecimentos',
                        'Ano': 2023
                    })

        # Ordenar por categoria e depois por valor
        all_fertilizer_data.sort(key=lambda x: (x['Categoria'], -x['Valor']))

        # Criar DataFrame principal
        df_main = pd.DataFrame(all_fertilizer_data)

        # Criar resumo por categoria
        category_summary = df_main.groupby('Categoria').agg({
            'Valor': ['sum', 'count', 'mean', 'max', 'min']
        }).round(2)
        category_summary.columns = ['Valor Total', 'Nº Municípios', 'Valor Médio', 'Valor Máximo', 'Valor Mínimo']
        category_summary = category_summary.sort_values('Valor Total', ascending=False)
        category_summary.reset_index(inplace=True)

        # Criar resumo por estado
        state_summary = df_main.groupby('UF').agg({
            'Valor': ['sum', 'count', 'mean']
        }).round(2)
        state_summary.columns = ['Valor Total', 'Nº Municípios', 'Valor Médio']
        state_summary = state_summary.sort_values('Valor Total', ascending=False)
        state_summary.reset_index(inplace=True)

        # Criar dados de resumo geral
        total_categories = df_main['Categoria'].nunique()
        total_municipalities = df_main['Código IBGE'].nunique()
        total_records = len(df_main)
        total_value = df_main['Valor'].sum()
        avg_value = df_main['Valor'].mean()

        general_summary = pd.DataFrame([
            ['Estatística', 'Valor'],
            ['Base de Dados', 'Fertilizantes - Censo Agropecuário 2017'],
            ['Ano de Referência', 2023],
            ['Total de Categorias', total_categories],
            ['Total de Municípios', total_municipalities],
            ['Total de Registros', total_records],
            ['Valor Total Geral', f'{total_value:,.0f}'],
            ['Valor Médio Geral', f'{avg_value:,.2f}'],
            ['Data da Exportação', pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')]
        ])

        # Criar arquivo Excel
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Planilha principal com todos os dados
            df_main.to_excel(writer, sheet_name='Dados Completos', index=False)

            # Planilha de resumo geral
            general_summary.to_excel(writer, sheet_name='Resumo Geral', index=False, header=False)

            # Planilha de resumo por categoria
            category_summary.to_excel(writer, sheet_name='Resumo por Categoria', index=False)

            # Planilha de resumo por estado
            state_summary.to_excel(writer, sheet_name='Resumo por Estado', index=False)

            # Criar planilhas separadas para cada categoria (máximo 10 categorias principais)
            top_categories = category_summary.head(10)
            for _, category_row in top_categories.iterrows():
                category_name = category_row['Categoria']
                safe_name = category_name.replace('/', '_').replace('\\', '_').replace(':', '_')[:30]

                category_df = df_main[df_main['Categoria'] == category_name].copy()
                category_df = category_df.sort_values('Valor', ascending=False)

                try:
                    category_df.to_excel(writer, sheet_name=safe_name, index=False)
                except Exception as e:
                    print(f"Erro ao criar planilha para categoria {category_name}: {e}")

        output.seek(0)

        # Nome do arquivo
        filename = f'base_completa_fertilizantes_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar base completa de fertilizantes: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/fertilizer-analysis/<category_name>')
def export_fertilizer_analysis(category_name):
    """Export fertilizer analysis data as Excel file"""
    try:
        # Obter parâmetro de estado opcional
        state_filter = request.args.get('state')

        if category_name not in FERTILIZER_DATA:
            return jsonify({'success': False, 'error': 'Categoria de fertilizantes não encontrada'}), 404

        # Preparar dados para exportação
        analysis_data = []
        for municipality_code, municipality_data in FERTILIZER_DATA[category_name].items():
            # Filtrar apenas municípios válidos
            municipality_code_str = str(municipality_code)
            municipality_name = municipality_data.get('municipality_name', '').lower()

            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and
                municipality_data.get('municipality_name') and
                not any(keyword in municipality_name for keyword in [
                    'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                    'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                    'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                    'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                ])):

                # Aplicar filtro de estado se especificado
                if state_filter and municipality_data.get('state_code') != state_filter:
                    continue

                analysis_data.append({
                    'Código IBGE': municipality_code,
                    'Município': municipality_data.get('municipality_name', 'Desconhecido'),
                    'UF': municipality_data.get('state_code', 'XX'),
                    'Categoria': category_name,
                    'Valor': municipality_data.get('value', 0),
                    'Unidade': 'estabelecimentos',
                    'Ano': 2023
                })

        # Ordenar por valor (maior para menor)
        analysis_data.sort(key=lambda x: x['Valor'], reverse=True)

        # Criar DataFrame
        df = pd.DataFrame(analysis_data)

        # Calcular estatísticas resumidas
        total_value = df['Valor'].sum()
        total_municipalities = len(df)
        average_value = df['Valor'].mean()
        max_value = df['Valor'].max()
        min_value = df['Valor'].min()

        # Criar dados de resumo estatístico
        summary_data = [
            ['Estatística', 'Valor'],
            ['Categoria Analisada', category_name],
            ['Filtro de Estado', state_filter if state_filter else 'Nacional (Todos os Estados)'],
            ['Ano de Referência', 2023],
            ['Total de Municípios', total_municipalities],
            ['Valor Total', f'{total_value:,.0f}'],
            ['Valor Médio por Município', f'{average_value:,.2f}'],
            ['Maior Valor Municipal', f'{max_value:,.0f}'],
            ['Menor Valor Municipal', f'{min_value:,.0f}'],
            ['Data da Exportação', pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')]
        ]

        # Criar resumo por estado
        state_summary = df.groupby('UF').agg({
            'Valor': ['sum', 'count', 'mean']
        }).round(2)
        state_summary.columns = ['Valor Total', 'Nº Municípios', 'Valor Médio']
        state_summary = state_summary.sort_values('Valor Total', ascending=False)
        state_summary.reset_index(inplace=True)

        # Criar arquivo Excel
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Planilha principal com dados detalhados
            df.to_excel(writer, sheet_name='Dados Detalhados', index=False)

            # Planilha de resumo estatístico
            summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            summary_df.to_excel(writer, sheet_name='Resumo Estatístico', index=False)

            # Planilha de resumo por estado
            state_summary.to_excel(writer, sheet_name='Resumo por Estado', index=False)

            # Top 20 maiores estabelecimentos
            top_20 = df.head(20).copy()
            top_20['Ranking'] = range(1, len(top_20) + 1)
            top_20 = top_20[['Ranking', 'Município', 'UF', 'Valor', 'Unidade']]
            top_20.to_excel(writer, sheet_name='Top 20', index=False)

        output.seek(0)

        # Nome do arquivo
        safe_category_name = category_name.replace('/', '_').replace('\\', '_').replace(':', '_')
        state_suffix = f'_{state_filter}' if state_filter else '_Nacional'
        filename = f'analise_{safe_category_name}{state_suffix}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar análise de fertilizantes: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/crop-analysis/<crop_name>')
def export_crop_analysis(crop_name):
    """Export crop analysis data as Excel file"""
    try:
        # Obter parâmetro de estado opcional
        state_filter = request.args.get('state')

        if crop_name not in CROP_DATA:
            return jsonify({'success': False, 'error': 'Cultura não encontrada'}), 404

        # Preparar dados para exportação
        analysis_data = []
        for municipality_code, municipality_data in CROP_DATA[crop_name].items():
            # Filtrar apenas municípios válidos
            municipality_code_str = str(municipality_code)
            municipality_name = municipality_data.get('municipality_name', '').lower()

            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and
                municipality_data.get('municipality_name') and
                not any(keyword in municipality_name for keyword in [
                    'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                    'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                    'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                    'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                ]) and
                municipality_name not in [
                    'alto teles pires', 'sudeste mato-grossense', 'parecis', 'barreiras',
                    'dourados', 'norte mato-grossense', 'portal da amazônia'
                ]):

                # Aplicar filtro de estado se especificado
                if state_filter and municipality_data.get('state_code') != state_filter:
                    continue

                analysis_data.append({
                    'Código IBGE': municipality_code,
                    'Município': municipality_data.get('municipality_name', 'Desconhecido'),
                    'UF': municipality_data.get('state_code', 'XX'),
                    'Cultura': crop_name,
                    'Área Colhida (hectares)': municipality_data.get('harvested_area', 0),
                    'Ano': 2023
                })

        # Ordenar por área colhida (maior para menor)
        analysis_data.sort(key=lambda x: x['Área Colhida (hectares)'], reverse=True)

        # Criar DataFrame
        df = pd.DataFrame(analysis_data)

        # Calcular estatísticas resumidas
        total_area = df['Área Colhida (hectares)'].sum()
        total_municipalities = len(df)
        average_area = df['Área Colhida (hectares)'].mean()
        max_area = df['Área Colhida (hectares)'].max()
        min_area = df['Área Colhida (hectares)'].min()

        # Criar dados de resumo estatístico
        summary_data = [
            ['Estatística', 'Valor'],
            ['Cultura Analisada', crop_name],
            ['Filtro de Estado', state_filter if state_filter else 'Nacional (Todos os Estados)'],
            ['Ano de Referência', 2023],
            ['Total de Municípios', total_municipalities],
            ['Área Total Colhida (ha)', f'{total_area:,.2f}'],
            ['Área Média por Município (ha)', f'{average_area:,.2f}'],
            ['Maior Área Municipal (ha)', f'{max_area:,.2f}'],
            ['Menor Área Municipal (ha)', f'{min_area:,.2f}'],
            ['Data da Exportação', pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')]
        ]

        # Criar resumo por estado
        state_summary = df.groupby('UF').agg({
            'Área Colhida (hectares)': ['sum', 'count', 'mean']
        }).round(2)
        state_summary.columns = ['Área Total (ha)', 'Nº Municípios', 'Área Média (ha)']
        state_summary = state_summary.sort_values('Área Total (ha)', ascending=False)
        state_summary.reset_index(inplace=True)

        # Criar arquivo Excel
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Planilha principal com dados detalhados
            df.to_excel(writer, sheet_name='Dados Detalhados', index=False)

            # Planilha de resumo estatístico
            summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            summary_df.to_excel(writer, sheet_name='Resumo Estatístico', index=False)

            # Planilha de resumo por estado
            state_summary.to_excel(writer, sheet_name='Resumo por Estado', index=False)

            # Top 20 maiores produtores
            top_20 = df.head(20).copy()
            top_20['Ranking'] = range(1, len(top_20) + 1)
            top_20 = top_20[['Ranking', 'Município', 'UF', 'Área Colhida (hectares)']]
            top_20.to_excel(writer, sheet_name='Top 20 Produtores', index=False)

        output.seek(0)

        # Nome do arquivo
        safe_crop_name = crop_name.replace('/', '_').replace('\\', '_').replace(':', '_')
        state_suffix = f'_{state_filter}' if state_filter else '_Nacional'
        filename = f'analise_{safe_crop_name}{state_suffix}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar análise: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/agrotoxico-analysis/<category>')
def export_agrotoxico_analysis(category):
    """Export agrotóxico analysis data as Excel file"""
    try:
        state_filter = request.args.get('state')

        if category not in AGROTOXICO_DATA:
            return jsonify({'success': False, 'error': 'Categoria de agrotóxico não encontrada'}), 404

        analysis_data = []
        for municipality_code, municipality_data in AGROTOXICO_DATA[category].items():
            municipality_code_str = str(municipality_code)
            municipality_name = municipality_data.get('municipality_name', '').lower()

            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and
                municipality_data.get('municipality_name')):

                if state_filter and municipality_data.get('state_code') != state_filter:
                    continue

                analysis_data.append({
                    'Código IBGE': municipality_code,
                    'Município': municipality_data.get('municipality_name', 'Desconhecido'),
                    'UF': municipality_data.get('state_code', 'XX'),
                    'Categoria': category,
                    'Valor': municipality_data.get('value', 0),
                    'Unidade': municipality_data.get('unit', 'un'),
                    'Ano': 2023
                })

        analysis_data.sort(key=lambda x: x['Valor'], reverse=True)
        df = pd.DataFrame(analysis_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Dados Detalhados', index=False)

        output.seek(0)
        safe_category_name = category.replace('/', '_').replace('\\', '_').replace(':', '_')
        state_suffix = f'_{state_filter}' if state_filter else '_Nacional'
        filename = f'analise_agrotoxico_{safe_category_name}{state_suffix}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar análise de agrotóxico: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/consultoria-analysis/<category>')
def export_consultoria_analysis(category):
    """Export consultoria analysis data as Excel file"""
    try:
        state_filter = request.args.get('state')

        if category not in CONSULTORIA_DATA:
            return jsonify({'success': False, 'error': 'Categoria de consultoria não encontrada'}), 404

        analysis_data = []
        for municipality_code, municipality_data in CONSULTORIA_DATA[category].items():
            municipality_code_str = str(municipality_code)

            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and
                municipality_data.get('municipality_name')):

                if state_filter and municipality_data.get('state_code') != state_filter:
                    continue

                analysis_data.append({
                    'Código IBGE': municipality_code,
                    'Município': municipality_data.get('municipality_name', 'Desconhecido'),
                    'UF': municipality_data.get('state_code', 'XX'),
                    'Categoria': category,
                    'Valor': municipality_data.get('value', 0),
                    'Unidade': municipality_data.get('unit', 'un'),
                    'Ano': 2023
                })

        analysis_data.sort(key=lambda x: x['Valor'], reverse=True)
        df = pd.DataFrame(analysis_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Dados Detalhados', index=False)

        output.seek(0)
        safe_category_name = category.replace('/', '_').replace('\\', '_').replace(':', '_')
        state_suffix = f'_{state_filter}' if state_filter else '_Nacional'
        filename = f'analise_consultoria_{safe_category_name}{state_suffix}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar análise de consultoria: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/corretivos-analysis/<category>')
def export_corretivos_analysis(category):
    """Export corretivos analysis data as Excel file"""
    try:
        state_filter = request.args.get('state')

        if category not in CORRETIVOS_DATA:
            return jsonify({'success': False, 'error': 'Categoria de corretivo não encontrada'}), 404

        analysis_data = []
        for municipality_code, municipality_data in CORRETIVOS_DATA[category].items():
            municipality_code_str = str(municipality_code)

            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and
                municipality_data.get('municipality_name')):

                if state_filter and municipality_data.get('state_code') != state_filter:
                    continue

                analysis_data.append({
                    'Código IBGE': municipality_code,
                    'Município': municipality_data.get('municipality_name', 'Desconhecido'),
                    'UF': municipality_data.get('state_code', 'XX'),
                    'Categoria': category,
                    'Valor': municipality_data.get('value', 0),
                    'Unidade': municipality_data.get('unit', 'un'),
                    'Ano': 2023
                })

        analysis_data.sort(key=lambda x: x['Valor'], reverse=True)
        df = pd.DataFrame(analysis_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Dados Detalhados', index=False)

        output.seek(0)
        safe_category_name = category.replace('/', '_').replace('\\', '_').replace(':', '_')
        state_suffix = f'_{state_filter}' if state_filter else '_Nacional'
        filename = f'analise_corretivo_{safe_category_name}{state_suffix}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar análise de corretivo: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/despesa-analysis/<category>')
def export_despesa_analysis(category):
    """Export despesa analysis data as Excel file"""
    try:
        state_filter = request.args.get('state')

        if category not in DESPESA_DATA:
            return jsonify({'success': False, 'error': 'Categoria de despesa não encontrada'}), 404

        analysis_data = []
        for municipality_code, municipality_data in DESPESA_DATA[category].items():
            municipality_code_str = str(municipality_code)

            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and
                municipality_data.get('municipality_name')):

                if state_filter and municipality_data.get('state_code') != state_filter:
                    continue

                analysis_data.append({
                    'Código IBGE': municipality_code,
                    'Município': municipality_data.get('municipality_name', 'Desconhecido'),
                    'UF': municipality_data.get('state_code', 'XX'),
                    'Categoria': category,
                    'Valor': municipality_data.get('value', 0),
                    'Unidade': municipality_data.get('unit', 'R$'),
                    'Ano': 2023
                })

        analysis_data.sort(key=lambda x: x['Valor'], reverse=True)
        df = pd.DataFrame(analysis_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Dados Detalhados', index=False)

        output.seek(0)
        safe_category_name = category.replace('/', '_').replace('\\', '_').replace(':', '_')
        state_suffix = f'_{state_filter}' if state_filter else '_Nacional'
        filename = f'analise_despesa_{safe_category_name}{state_suffix}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar análise de despesa: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/escolaridade-analysis/<category>')
def export_escolaridade_analysis(category):
    """Export escolaridade analysis data as Excel file"""
    try:
        state_filter = request.args.get('state')

        if category not in ESCOLARIDADE_DATA:
            return jsonify({'success': False, 'error': 'Categoria de escolaridade não encontrada'}), 404

        analysis_data = []
        for municipality_code, municipality_data in ESCOLARIDADE_DATA[category].items():
            municipality_code_str = str(municipality_code)

            if (len(municipality_code_str) == 7 and 
                municipality_code_str.isdigit() and
                municipality_code_str[0] in '12345' and
                municipality_data.get('municipality_name')):

                if state_filter and municipality_data.get('state_code') != state_filter:
                    continue

                analysis_data.append({
                    'Código IBGE': municipality_code,
                    'Município': municipality_data.get('municipality_name', 'Desconhecido'),
                    'UF': municipality_data.get('state_code', 'XX'),
                    'Categoria': category,
                    'Valor': municipality_data.get('value', 0),
                    'Unidade': municipality_data.get('unit', 'un'),
                    'Ano': 2023
                })

        analysis_data.sort(key=lambda x: x['Valor'], reverse=True)
        df = pd.DataFrame(analysis_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Dados Detalhados', index=False)

        output.seek(0)
        safe_category_name = category.replace('/', '_').replace('\\', '_').replace(':', '_')
        state_suffix = f'_{state_filter}' if state_filter else '_Nacional'
        filename = f'analise_escolaridade_{safe_category_name}{state_suffix}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar análise de escolaridade: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/analise-comercial/excel/<int:revenda_id>')
@login_required
def export_revenda_commercial_analysis(revenda_id):
    """Exportar análise comercial completa da revenda em Excel"""
    try:
        revenda = Revenda.query.get_or_404(revenda_id)
        municipios_codes = revenda.get_municipios_list()

        if not municipios_codes:
            return jsonify({'success': False, 'error': 'Nenhum município encontrado para esta revenda'})

        # Obter dados completos da análise
        analysis_data = calculate_revenda_analysis(municipios_codes)

        # Criar arquivo Excel
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Planilha de Resumo Geral
            resumo_geral_data = [
                ['Informações da Revenda', ''],
                ['Nome', revenda.nome],
                ['CNPJ', revenda.cnpj],
                ['CNAE', revenda.cnae],
                ['Cor', revenda.cor],
                ['Total de Municípios', len(municipios_codes)],
                ['Data da Análise', pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')],
                ['', ''],
                ['Resumo Financeiro', ''],
                ['Total Receita (R$)', f'{analysis_data["financialData"]["totalReceita"]:,.2f}'],
                ['Total Despesa (R$)', f'{analysis_data["financialData"]["totalDespesa"]:,.2f}'],
                ['Saldo (R$)', f'{analysis_data["financialData"]["saldo"]:,.2f}'],
                ['', ''],
                ['Resumo Agrícola', ''],
                ['Total de Culturas', len(analysis_data['cropsData']['crops'])],
                ['Área Total Cultivada (ha)', f'{sum([c["total_area"] for c in analysis_data["cropsData"]["crops"]]):,.2f}'],
                ['', ''],
                ['Resumo de Insumos', ''],
                ['Categorias de Fertilizantes', len(analysis_data['fertilizerData']['categories'])],
                ['Categorias de Agrotóxicos', len(analysis_data['agrotoxicoData']['categories'])],
                ['Categorias de Consultoria', len(analysis_data['consultoriaData']['categories'])],
                ['Categorias de Corretivos', len(analysis_data['corretivosData']['categories'])],
                ['Categorias de Escolaridade', len(analysis_data['escolaridadeData']['categories'])]
            ]

            resumo_df = pd.DataFrame(resumo_geral_data, columns=['Item', 'Valor'])
            resumo_df.to_excel(writer, sheet_name='Resumo Geral', index=False)

            # 2. Planilha Financeira Detalhada
            financial_detail_data = []
            for municipio in analysis_data['financialData']['municipios']:
                financial_detail_data.append({
                    'Código IBGE': municipio['code'],
                    'Município': municipio['name'],
                    'UF': municipio['state'],
                    'Receita (R$)': municipio['receita'],
                    'Despesa (R$)': municipio['despesa'],
                    'Saldo (R$)': municipio['saldo'],
                    'Margem (%)': (municipio['saldo'] / max(municipio['receita'], 1) * 100) if municipio['receita'] > 0 else 0
                })

            if financial_detail_data:
                financial_df = pd.DataFrame(financial_detail_data)
                financial_df = financial_df.sort_values('Saldo (R$)', ascending=False)
                financial_df.to_excel(writer, sheet_name='Análise Financeira', index=False)

            # 3. Planilha de Culturas Detalhada
            crops_detail_data = []
            total_crop_area = sum([c['total_area'] for c in analysis_data['cropsData']['crops']])

            for crop in analysis_data['cropsData']['crops']:
                participation = (crop['total_area'] / total_crop_area * 100) if total_crop_area > 0 else 0
                avg_area_per_municipality = crop['total_area'] / max(crop['municipalities_count'], 1)

                crops_detail_data.append({
                    'Cultura': crop['name'],
                    'Área Total (ha)': crop['total_area'],
                    'Nº Municípios': crop['municipalities_count'],
                    'Área Média por Município (ha)': avg_area_per_municipality,
                    'Participação (%)': participation,
                    'Ranking': crops_detail_data.__len__() + 1
                })

            if crops_detail_data:
                crops_df = pd.DataFrame(crops_detail_data)
                crops_df.to_excel(writer, sheet_name='Culturas Detalhado', index=False)

            # 4. Planilha de Fertilizantes Detalhada
            fertilizer_detail_data = []
            total_fertilizer = sum([c['total'] for c in analysis_data['fertilizerData']['categories']])

            for category in analysis_data['fertilizerData']['categories']:
                participation = (category['total'] / total_fertilizer * 100) if total_fertilizer > 0 else 0

                fertilizer_detail_data.append({
                    'Categoria': category['name'],
                    'Total Estabelecimentos': category['total'],
                    'Participação (%)': participation,
                    'Ranking': len(fertilizer_detail_data) + 1
                })

            if fertilizer_detail_data:
                fertilizer_df = pd.DataFrame(fertilizer_detail_data)
                fertilizer_df.to_excel(writer, sheet_name='Fertilizantes', index=False)

            # 5. Planilha de Agrotóxicos Detalhada
            agrotoxico_detail_data = []
            total_agrotoxico = sum([c['total'] for c in analysis_data['agrotoxicoData']['categories']])

            for category in analysis_data['agrotoxicoData']['categories']:
                participation = (category['total'] / total_agrotoxico * 100) if total_agrotoxico > 0 else 0

                agrotoxico_detail_data.append({
                    'Categoria': category['name'],
                    'Total Estabelecimentos': category['total'],
                    'Participação (%)': participation,
                    'Ranking': len(agrotoxico_detail_data) + 1
                })

            if agrotoxico_detail_data:
                agrotoxico_df = pd.DataFrame(agrotoxico_detail_data)
                agrotoxico_df.to_excel(writer, sheet_name='Agrotóxicos', index=False)

            # 6. Planilha de Consultoria Detalhada
            consultoria_detail_data = []
            total_consultoria = sum([c['total'] for c in analysis_data['consultoriaData']['categories']])

            for category in analysis_data['consultoriaData']['categories']:
                participation = (category['total'] / total_consultoria * 100) if total_consultoria > 0 else 0

                consultoria_detail_data.append({
                    'Categoria': category['name'],
                    'Total Estabelecimentos': category['total'],
                    'Participação (%)': participation,
                    'Ranking': len(consultoria_detail_data) + 1
                })

            if consultoria_detail_data:
                consultoria_df = pd.DataFrame(consultoria_detail_data)
                consultoria_df.to_excel(writer, sheet_name='Consultoria Técnica', index=False)

            # 7. Planilha de Corretivos Detalhada
            corretivos_detail_data = []
            total_corretivos = sum([c['total'] for c in analysis_data['corretivosData']['categories']])

            for category in analysis_data['corretivosData']['categories']:
                participation = (category['total'] / total_corretivos * 100) if total_corretivos > 0 else 0

                corretivos_detail_data.append({
                    'Categoria': category['name'],
                    'Total Estabelecimentos': category['total'],
                    'Participação (%)': participation,
                    'Ranking': len(corretivos_detail_data) + 1
                })

            if corretivos_detail_data:
                corretivos_df = pd.DataFrame(corretivos_detail_data)
                corretivos_df.to_excel(writer, sheet_name='Corretivos', index=False)

            # 8. Planilha de Escolaridade Detalhada
            escolaridade_detail_data = []
            total_escolaridade = sum([c['total'] for c in analysis_data['escolaridadeData']['categories']])

            for category in analysis_data['escolaridadeData']['categories']:
                participation = (category['total'] / total_escolaridade * 100) if total_escolaridade > 0 else 0

                escolaridade_detail_data.append({
                    'Categoria': category['name'],
                    'Total Pessoas': category['total'],
                    'Participação (%)': participation,
                    'Ranking': len(escolaridade_detail_data) + 1
                })

            if escolaridade_detail_data:
                escolaridade_df = pd.DataFrame(escolaridade_detail_data)
                escolaridade_df.to_excel(writer, sheet_name='Escolaridade', index=False)

            # 9. Planilha de Municípios da Revenda
            municipios_detail_data = []

            # Buscar dados dos municípios
            for municipio_code in municipios_codes:
                municipio_code_str = str(municipio_code)
                municipio_name = ""
                state_code = ""

                # Buscar dados em qualquer fonte para obter nome e estado
                for crop_data_entry in CROP_DATA.values():
                    if municipio_code_str in crop_data_entry:
                        municipio_data = crop_data_entry[municipio_code_str]
                        municipio_name = municipio_data.get('municipality_name', '')
                        state_code = municipio_data.get('state_code', '')
                        break

                if notmunicipio_name:
                    municipio_name = f"Município {municipio_code}"
                    state_code = "XX"

                municipios_detail_data.append({
                    'Código IBGE': municipio_code,
                    'Nome do Município': municipio_name,
                    'UF': state_code
                })

            municipios_df = pd.DataFrame(municipios_detail_data)
            municipios_df = municipios_df.sort_values(['UF', 'Nome do Município'])
            municipios_df.to_excel(writer, sheet_name='Municípios da Revenda', index=False)

        output.seek(0)

        # Nome do arquivo
        safe_name = revenda.nome.replace(' ', '_').replace('/', '_').replace('\\', '_')
        filename = f'analise_comercial_{safe_name}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar análise comercial da revenda: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Rota duplicada removida aqui para evitar o erro AssertionError
# @app.route('/api/analise-comercial-vendedor/excel/<int:vendedor_id>')
# @login_required
# def export_vendedor_commercial_analysis(vendedor_id):
#     """Exportar análise comercial completa do vendedor em Excel"""
#     try:
#         vendedor = Vendedor.query.get_or_404(vendedor_id)
#         municipios_codes = vendedor.get_municipios_list()

#         if not municipios_codes:
#             return jsonify({'success': False, 'error': 'Nenhum município encontrado para este vendedor'})

#         # Obter dados completos da análise
#         analysis_data = calculate_revenda_analysis(municipios_codes)

#         # Criar arquivo Excel
#         output = io.BytesIO()

#         with pd.ExcelWriter(output, engine='openpyxl') as writer:
#             # 1. Planilha de Resumo Geral
#             resumo_geral_data = [
#                 ['Informações do Vendedor', ''],
#                 ['Nome', vendedor.nome],
#                 ['E-mail', vendedor.email],
#                 ['Telefone', vendedor.telefone],
#                 ['CPF', vendedor.cpf],
#                 ['Cor', vendedor.cor],
#                 ['Total de Municípios', len(municipios_codes)],
#                 ['Data da Análise', pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')],
#                 ['', ''],
#                 ['Resumo Financeiro', ''],
#                 ['Total Receita (R$)', f'{analysis_data["financialData"]["totalReceita"]:,.2f}'],
#                 ['Total Despesa (R$)', f'{analysis_data["financialData"]["totalDespesa"]:,.2f}'],
#                 ['Saldo (R$)', f'{analysis_data["financialData"]["saldo"]:,.2f}'],
#                 ['', ''],
#                 ['Resumo Agrícola', ''],
#                 ['Total de Culturas', len(analysis_data['cropsData']['crops'])],
#                 ['Área Total Cultivada (ha)', f'{sum([c["total_area"] for c in analysis_data["cropsData"]["crops"]]):,.2f}'],
#                 ['', ''],
#                 ['Resumo de Insumos', ''],
#                 ['Categorias de Fertilizantes', len(analysis_data['fertilizerData']['categories'])],
#                 ['Categorias de Agrotóxicos', len(analysis_data['agrotoxicoData']['categories'])],
#                 ['Categorias de Consultoria', len(analysis_data['consultoriaData']['categories'])],
#                 ['Categorias de Corretivos', len(analysis_data['corretivosData']['categories'])],
#                 ['Categorias de Escolaridade', len(analysis_data['escolaridadeData']['categories'])]
#             ]

#             resumo_df = pd.DataFrame(resumo_geral_data, columns=['Item', 'Valor'])
#             resumo_df.to_excel(writer, sheet_name='Resumo Geral', index=False)

#             # 2. Planilha Financeira Detalhada
#             financial_detail_data = []
#             for municipio in analysis_data['financialData']['municipios']:
#                 financial_detail_data.append({
#                     'Código IBGE': municipio['code'],
#                     'Município': municipio['name'],
#                     'UF': municipio['state'],
#                     'Receita (R$)': municipio['receita'],
#                     'Despesa (R$)': municipio['despesa'],
#                     'Saldo (R$)': municipio['saldo'],
#                     'Margem (%)': (municipio['saldo'] / max(municipio['receita'], 1) * 100) if municipio['receita'] > 0 else 0
#                 })

#             if financial_detail_data:
#                 financial_df = pd.DataFrame(financial_detail_data)
#                 financial_df = financial_df.sort_values('Saldo (R$)', ascending=False)
#                 financial_df.to_excel(writer, sheet_name='Análise Financeira', index=False)

#             # 3. Planilha de Culturas Detalhada
#             crops_detail_data = []
#             total_crop_area = sum([c['total_area'] for c in analysis_data['cropsData']['crops']])

#             for i, crop in enumerate(analysis_data['cropsData']['crops']):
#                 participation = (crop['total_area'] / total_crop_area * 100) if total_crop_area > 0 else 0
#                 avg_area_per_municipality = crop['total_area'] / max(crop['municipalities_count'], 1)

#                 crops_detail_data.append({
#                     'Ranking': i + 1,
#                     'Cultura': crop['name'],
#                     'Área Total (ha)': crop['total_area'],
#                     'Nº Municípios': crop['municipalities_count'],
#                     'Área Média por Município (ha)': avg_area_per_municipality,
#                     'Participação (%)': participation
#                 })

#             if crops_detail_data:
#                 crops_df = pd.DataFrame(crops_detail_data)
#                 crops_df.to_excel(writer, sheet_name='Culturas Detalhado', index=False)

#             # 4. Planilha de Fertilizantes Detalhada
#             fertilizer_detail_data = []
#             total_fertilizer = sum([c['total'] for c in analysis_data['fertilizerData']['categories']])

#             for i, category in enumerate(analysis_data['fertilizerData']['categories']):
#                 participation = (category['total'] / total_fertilizer * 100) if total_fertilizer > 0 else 0

#                 fertilizer_detail_data.append({
#                     'Ranking': i + 1,
#                     'Categoria': category['name'],
#                     'Total Estabelecimentos': category['total'],
#                     'Participação (%)': participation
#                 })

#             if fertilizer_detail_data:
#                 fertilizer_df = pd.DataFrame(fertilizer_detail_data)
#                 fertilizer_df.to_excel(writer, sheet_name='Fertilizantes', index=False)

#             # 5. Planilha de Agrotóxicos Detalhada
#             agrotoxico_detail_data = []
#             total_agrotoxico = sum([c['total'] for c in analysis_data['agrotoxicoData']['categories']])

#             for i, category in enumerate(analysis_data['agrotoxicoData']['categories']):
#                 participation = (category['total'] / total_agrotoxico * 100) if total_agrotoxico > 0 else 0

#                 agrotoxico_detail_data.append({
#                     'Ranking': i + 1,
#                     'Categoria': category['name'],
#                     'Total Estabelecimentos': category['total'],
#                     'Participação (%)': participation
#                 })

#             if agrotoxico_detail_data:
#                 agrotoxico_df = pd.DataFrame(agrotoxico_detail_data)
#                 agrotoxico_df.to_excel(writer, sheet_name='Agrotóxicos', index=False)

#             # 6. Planilha de Consultoria Detalhada
#             consultoria_detail_data = []
#             total_consultoria = sum([c['total'] for c in analysis_data['consultoriaData']['categories']])

#             for i, category in enumerate(analysis_data['consultoriaData']['categories']):
#                 participation = (category['total'] / total_consultoria * 100) if total_consultoria > 0 else 0

#                 consultoria_detail_data.append({
#                     'Ranking': i + 1,
#                     'Categoria': category['name'],
#                     'Total Estabelecimentos': category['total'],
#                     'Participação (%)': participation
#                 })

#             if consultoria_detail_data:
#                 consultoria_df = pd.DataFrame(consultoria_detail_data)
#                 consultoria_df.to_excel(writer, sheet_name='Consultoria Técnica', index=False)

#             # 7. Planilha de Corretivos Detalhada
#             corretivos_detail_data = []
#             total_corretivos = sum([c['total'] for c in analysis_data['corretivosData']['categories']])

#             for i, category in enumerate(analysis_data['corretivosData']['categories']):
#                 participation = (category['total'] / total_corretivos * 100) if total_corretivos > 0 else 0

#                 corretivos_detail_data.append({
#                     'Ranking': i + 1,
#                     'Categoria': category['name'],
#                     'Total Estabelecimentos': category['total'],
#                     'Participação (%)': participation
#                 })

#             if corretivos_detail_data:
#                 corretivos_df = pd.DataFrame(corretivos_detail_data)
#                 corretivos_df.to_excel(writer, sheet_name='Corretivos', index=False)

#             # 8. Planilha de Escolaridade Detalhada
#             escolaridade_detail_data = []
#             total_escolaridade = sum([c['total'] for c in analysis_data['escolaridadeData']['categories']])

#             for i, category in enumerate(analysis_data['escolaridadeData']['categories']):
#                 participation = (category['total'] / total_escolaridade * 100) if total_escolaridade > 0 else 0

#                 escolaridade_detail_data.append({
#                     'Ranking': i + 1,
#                     'Categoria': category['name'],
#                     'Total Pessoas': category['total'],
#                     'Participação (%)': participation
#                 })

#             if escolaridade_detail_data:
#                 escolaridade_df = pd.DataFrame(escolaridade_detail_data)
#                 escolaridade_df.to_excel(writer, sheet_name='Escolaridade', index=False)

#             # 9. Planilha de Municípios da Revenda
#             municipios_detail_data = []

#             # Buscar dados dos municípios
#             for municipio_code in municipios_codes:
#                 municipio_code_str = str(municipio_code)
#                 municipio_name = ""
#                 state_code = ""

#                 # Buscar dados em qualquer fonte para obter nome e estado
#                 for crop_data_entry in CROP_DATA.values():
#                     if municipio_code_str in crop_data_entry:
#                         municipio_data = crop_data_entry[municipio_code_str]
#                         municipio_name = municipio_data.get('municipality_name', '')
#                         state_code = municipio_data.get('state_code', '')
#                         break

#                 if not municipio_name:
#                     municipio_name = f"Município {municipio_code}"
#                     state_code = "XX"

#                 municipios_detail_data.append({
#                     'Código IBGE': municipio_code,
#                     'Nome do Município': municipio_name,
#                     'UF': state_code
#                 })

#             municipios_df = pd.DataFrame(municipios_detail_data)
#             municipios_df = municipios_df.sort_values(['UF', 'Nome do Município'])
#             municipios_df.to_excel(writer, sheet_name='Municípios da Revenda', index=False)

#         output.seek(0)

#         # Nome do arquivo
#         safe_name = revenda.nome.replace(' ', '_').replace('/', '_').replace('\\', '_')
#         filename = f'analise_comercial_{safe_name}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

#         return send_file(
#             output,
#             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#             as_attachment=True,
#             download_name=filename
#         )

#     except Exception as e:
#         print(f"Erro ao exportar análise comercial da revenda: {e}")
#         import traceback
#         traceback.print_exc()
#         return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/analise-comercial-vendedor/excel/<int:vendedor_id>')
@login_required
def export_vendedor_commercial_analysis(vendedor_id):
    """Exportar análise comercial completa do vendedor em Excel"""
    try:
        vendedor = Vendedor.query.get_or_404(vendedor_id)
        municipios_codes = vendedor.get_municipios_list()

        if not municipios_codes:
            return jsonify({'success': False, 'error': 'Nenhum município encontrado para este vendedor'})

        # Obter dados completos da análise
        analysis_data = calculate_revenda_analysis(municipios_codes)

        # Criar arquivo Excel
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Planilha de Resumo Geral
            resumo_geral_data = [
                ['Informações do Vendedor', ''],
                ['Nome', vendedor.nome],
                ['E-mail', vendedor.email],
                ['Telefone', vendedor.telefone],
                ['CPF', vendedor.cpf],
                ['Cor', vendedor.cor],
                ['Total de Municípios', len(municipios_codes)],
                ['Data da Análise', pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')],
                ['', ''],
                ['Resumo Financeiro', ''],
                ['Total Receita (R$)', f'{analysis_data["financialData"]["totalReceita"]:,.2f}'],
                ['Total Despesa (R$)', f'{analysis_data["financialData"]["totalDespesa"]:,.2f}'],
                ['Saldo (R$)', f'{analysis_data["financialData"]["saldo"]:,.2f}'],
                ['', ''],
                ['Resumo Agrícola', ''],
                ['Total de Culturas', len(analysis_data['cropsData']['crops'])],
                ['Área Total Cultivada (ha)', f'{sum([c["total_area"] for c in analysis_data["cropsData"]["crops"]]):,.2f}'],
                ['', ''],
                ['Resumo de Insumos', ''],
                ['Categorias de Fertilizantes', len(analysis_data['fertilizerData']['categories'])],
                ['Categorias de Agrotóxicos', len(analysis_data['agrotoxicoData']['categories'])],
                ['Categorias de Consultoria', len(analysis_data['consultoriaData']['categories'])],
                ['Categorias de Corretivos', len(analysis_data['corretivosData']['categories'])],
                ['Categorias de Escolaridade', len(analysis_data['escolaridadeData']['categories'])]
            ]

            resumo_df = pd.DataFrame(resumo_geral_data, columns=['Item', 'Valor'])
            resumo_df.to_excel(writer, sheet_name='Resumo Geral', index=False)

            # 2. Planilha Financeira Detalhada
            financial_detail_data = []
            for municipio in analysis_data['financialData']['municipios']:
                financial_detail_data.append({
                    'Código IBGE': municipio['code'],
                    'Município': municipio['name'],
                    'UF': municipio['state'],
                    'Receita (R$)': municipio['receita'],
                    'Despesa (R$)': municipio['despesa'],
                    'Saldo (R$)': municipio['saldo'],
                    'Margem (%)': (municipio['saldo'] / max(municipio['receita'], 1) * 100) if municipio['receita'] > 0 else 0
                })

            if financial_detail_data:
                financial_df = pd.DataFrame(financial_detail_data)
                financial_df = financial_df.sort_values('Saldo (R$)', ascending=False)
                financial_df.to_excel(writer, sheet_name='Análise Financeira', index=False)

            # 3. Planilha de Culturas Detalhada
            crops_detail_data = []
            total_crop_area = sum([c['total_area'] for c in analysis_data['cropsData']['crops']])

            for i, crop in enumerate(analysis_data['cropsData']['crops']):
                participation = (crop['total_area'] / total_crop_area * 100) if total_crop_area > 0 else 0
                avg_area_per_municipality = crop['total_area'] / max(crop['municipalities_count'], 1)

                crops_detail_data.append({
                    'Ranking': i + 1,
                    'Cultura': crop['name'],
                    'Área Total (ha)': crop['total_area'],
                    'Nº Municípios': crop['municipalities_count'],
                    'Área Média por Município (ha)': avg_area_per_municipality,
                    'Participação (%)': participation
                })

            if crops_detail_data:
                crops_df = pd.DataFrame(crops_detail_data)
                crops_df.to_excel(writer, sheet_name='Culturas Detalhado', index=False)

            # 4. Planilha de Fertilizantes Detalhada
            fertilizer_detail_data = []
            total_fertilizer = sum([c['total'] for c in analysis_data['fertilizerData']['categories']])

            for i, category in enumerate(analysis_data['fertilizerData']['categories']):
                participation = (category['total'] / total_fertilizer * 100) if total_fertilizer > 0 else 0

                fertilizer_detail_data.append({
                    'Ranking': i + 1,
                    'Categoria': category['name'],
                    'Total Estabelecimentos': category['total'],
                    'Participação (%)': participation
                })

            if fertilizer_detail_data:
                fertilizer_df = pd.DataFrame(fertilizer_detail_data)
                fertilizer_df.to_excel(writer, sheet_name='Fertilizantes', index=False)

            # 5. Planilha de Agrotóxicos Detalhada
            agrotoxico_detail_data = []
            total_agrotoxico = sum([c['total'] for c in analysis_data['agrotoxicoData']['categories']])

            for i, category in enumerate(analysis_data['agrotoxicoData']['categories']):
                participation = (category['total'] / total_agrotoxico * 100) if total_agrotoxico > 0 else 0

                agrotoxico_detail_data.append({
                    'Ranking': i + 1,
                    'Categoria': category['name'],
                    'Total Estabelecimentos': category['total'],
                    'Participação (%)': participation
                })

            if agrotoxico_detail_data:
                agrotoxico_df = pd.DataFrame(agrotoxico_detail_data)
                agrotoxico_df.to_excel(writer, sheet_name='Agrotóxicos', index=False)

            # 6. Planilha de Consultoria Detalhada
            consultoria_detail_data = []
            total_consultoria = sum([c['total'] for c in analysis_data['consultoriaData']['categories']])

            for i, category in enumerate(analysis_data['consultoriaData']['categories']):
                participation = (category['total'] / total_consultoria * 100) if total_consultoria > 0 else 0

                consultoria_detail_data.append({
                    'Ranking': i + 1,
                    'Categoria': category['name'],
                    'Total Estabelecimentos': category['total'],
                    'Participação (%)': participation
                })

            if consultoria_detail_data:
                consultoria_df = pd.DataFrame(consultoria_detail_data)
                consultoria_df.to_excel(writer, sheet_name='Consultoria Técnica', index=False)

            # 7. Planilha de Corretivos Detalhada
            corretivos_detail_data = []
            total_corretivos = sum([c['total'] for c in analysis_data['corretivosData']['categories']])

            for i, category in enumerate(analysis_data['corretivosData']['categories']):
                participation = (category['total'] / total_corretivos * 100) if total_corretivos > 0 else 0

                corretivos_detail_data.append({
                    'Ranking': i + 1,
                    'Categoria': category['name'],
                    'Total Estabelecimentos': category['total'],
                    'Participação (%)': participation
                })

            if corretivos_detail_data:
                corretivos_df = pd.DataFrame(corretivos_detail_data)
                corretivos_df.to_excel(writer, sheet_name='Corretivos', index=False)

            # 8. Planilha de Escolaridade Detalhada
            escolaridade_detail_data = []
            total_escolaridade = sum([c['total'] for c in analysis_data['escolaridadeData']['categories']])

            for i, category in enumerate(analysis_data['escolaridadeData']['categories']):
                participation = (category['total'] / total_escolaridade * 100) if total_escolaridade > 0 else 0

                escolaridade_detail_data.append({
                    'Ranking': i + 1,
                    'Categoria': category['name'],
                    'Total Pessoas': category['total'],
                    'Participação (%)': participation
                })

            if escolaridade_detail_data:
                escolaridade_df = pd.DataFrame(escolaridade_detail_data)
                escolaridade_df.to_excel(writer, sheet_name='Escolaridade', index=False)

            # 9. Planilha de Municípios da Revenda
            municipios_detail_data = []

            # Buscar dados dos municípios
            for municipio_code in municipios_codes:
                municipio_code_str = str(municipio_code)
                municipio_name = ""
                state_code = ""

                # Buscar dados em qualquer fonte para obter nome e estado
                for crop_data_entry in CROP_DATA.values():
                    if municipio_code_str in crop_data_entry:
                        municipio_data = crop_data_entry[municipio_code_str]
                        municipio_name = municipio_data.get('municipality_name', '')
                        state_code = municipio_data.get('state_code', '')
                        break

                if not municipio_name:
                    municipio_name = f"Município {municipio_code}"
                    state_code = "XX"

                municipios_detail_data.append({
                    'Código IBGE': municipio_code,
                    'Nome do Município': municipio_name,
                    'UF': state_code
                })

            municipios_df = pd.DataFrame(municipios_detail_data)
            municipios_df = municipios_df.sort_values(['UF', 'Nome do Município'])
            municipios_df.to_excel(writer, sheet_name='Municípios da Revenda', index=False)

        output.seek(0)

        # Nome do arquivo
        safe_name = vendedor.nome.replace(' ', '_').replace('/', '_').replace('\\', '_')
        filename = f'analise_comercial_vendedor_{safe_name}_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar análise comercial da revenda: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Helper function to calculate analysis (updated to remove scoring)
def calculate_revenda_analysis(municipios_codes):
    analysis = {
        'financialData': {'municipios': [], 'totalReceita': 0, 'totalDespesa': 0, 'saldo': 0},
        'cropsData': {'crops': []},
        'fertilizerData': {'categories': []},
        'agrotoxicoData': {'categories': []},
        'consultoriaData': {'categories': []},
        'corretivosData': {'categories': []},
        'escolaridadeData': {'categories': []}
    }

    try:
        # Financial data
        for municipio_code in municipios_codes:
            municipio_code_str = str(municipio_code)
            receita = 0
            despesa = 0
            municipio_name = ""
            municipio_state = ""

            if 'Total' in RECEITA_DATA and municipio_code_str in RECEITA_DATA['Total']:
                receita_data = RECEITA_DATA['Total'][municipio_code_str]
                receita = receita_data.get('value', 0)
                municipio_name = receita_data.get('municipality_name', '')
                municipio_state = receita_data.get('state_code', '')

            if 'Total' in DESPESA_DATA and municipio_code_str in DESPESA_DATA['Total']:
                despesa_data = DESPESA_DATA['Total'][municipio_code_str]
                despesa = despesa_data.get('value', 0)
                if not municipio_name:
                    municipio_name = despesa_data.get('municipality_name', '')
                    municipio_state = despesa_data.get('state_code', '')

            if receita > 0 or despesa > 0:
                analysis['financialData']['municipios'].append({
                    'code': municipio_code, # Added for use in JS
                    'name': municipio_name,
                    'state': municipio_state,
                    'receita': receita,
                    'despesa': despesa,
                    'saldo': receita - despesa # Added for use in JS
                })
                analysis['financialData']['totalReceita'] += receita
                analysis['financialData']['totalDespesa'] += despesa

        analysis['financialData']['saldo'] = analysis['financialData']['totalReceita'] - analysis['financialData']['totalDespesa']

        # Crops data - get ALL cultures found in revenda
        crop_totals = {}
        for crop_name, crop_data in CROP_DATA.items():
            total_area = 0
            municipalities_count = 0

            for municipio_code in municipios_codes:
                if str(municipio_code) in crop_data:
                    total_area += crop_data[str(municipio_code)].get('harvested_area', 0)
                    municipalities_count += 1

            if total_area > 0:
                crop_totals[crop_name] = {
                    'name': crop_name,
                    'total_area': total_area,
                    'municipalities_count': municipalities_count
                }

        analysis['cropsData']['crops'] = sorted(crop_totals.values(), key=lambda x: x['total_area'], reverse=True)

        # Helper function for other data categories
        def process_category_data(data_dict, analysis_key):
            category_totals = {}
            for category_name, category_data in data_dict.items():
                # Skip specific keys that represent totals or metadata, not actual categories
                if category_name in ['Total Estabelecimentos', 'Total estabelecimentos', 'TOTAL']:
                    continue

                total = 0
                for municipio_code in municipios_codes:
                    if str(municipio_code) in category_data:
                        total += category_data[str(municipio_code)].get('value', 0)

                if total > 0:
                    category_totals[category_name] = {
                        'name': category_name,
                        'total': total
                    }

            analysis[analysis_key]['categories'] = sorted(category_totals.values(), key=lambda x: x['total'], reverse=True)

        # Process all category data
        process_category_data(FERTILIZER_DATA, 'fertilizerData')
        process_category_data(AGROTOXICO_DATA, 'agrotoxicoData')
        process_category_data(CONSULTORIA_DATA, 'consultoriaData')
        process_category_data(CORRETIVOS_DATA, 'corretivosData')
        process_category_data(ESCOLARIDADE_DATA, 'escolaridadeData')

    except Exception as e:
        print(f"Error in calculate_revenda_analysis: {e}")
        import traceback
        traceback.print_exc()

    return analysis

# Rotas de Autenticação
@app.route('/login')
def login_page():
    if auth_manager.is_authenticated():
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/register')
def register_page():
    if auth_manager.is_authenticated():
        return redirect(url_for('index'))
    return render_template('register.html')

@app.route('/api/auth/register', methods=['POST'])
def api_register():
    try:
        data = request.get_json()

        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '')
        full_name = data.get('full_name', '').strip()
        confirm_password = data.get('confirm_password', '')

        if password != confirm_password:
            return jsonify({'success': False, 'error': 'Senhas não coincidem'})

        result = auth_manager.register_user(username, email, password, full_name)
        return jsonify(result)

    except Exception as e:
        return jsonify({'success': False, 'error': f'Erro interno: {str(e)}'})

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    try:
        data = request.get_json()

        username = data.get('username', '').strip()
        password = data.get('password', '')

        if not username or not password:
            return jsonify({'success': False, 'error': 'Username e senha são obrigatórios'})

        result = auth_manager.login_user(username, password)
        return jsonify(result)

    except Exception as e:
        return jsonify({'success': False, 'error': f'Erro interno: {str(e)}'})

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    try:
        result = auth_manager.logout_user()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Erro interno: {str(e)}'})

@app.route('/api/auth/profile')
@login_required
def api_profile():
    try:
        user = auth_manager.get_current_user()
        return jsonify({'success': True, 'user': user})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Erro interno: {str(e)}'})

@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def api_change_password():
    try:
        data = request.get_json()
        username = session.get('username')

        old_password = data.get('old_password', '')
        new_password = data.get('new_password', '')
        confirm_password = data.get('confirm_password', '')

        if new_password != confirm_password:
            return jsonify({'success': False, 'error': 'Senhas não coincidem'})

        result = auth_manager.change_password(username, old_password, new_password)
        return jsonify(result)

    except Exception as e:
        return jsonify({'success': False, 'error': f'Erro interno: {str(e)}'})

@app.route('/logout')
def logout():
    auth_manager.logout_user()
    return redirect(url_for('login_page'))

@app.route('/profile')
@login_required
def profile_page():
    user = auth_manager.get_current_user()
    return render_template('profile.html', user=user)

# Rotas para Revendas
@app.route('/revendas')
@login_required
def revendas_page():
    user = auth_manager.get_current_user()
    return render_template('revendas.html', user=user)

@app.route('/vendedores')
@login_required
def vendedores_page():
    user = auth_manager.get_current_user()
    return render_template('vendedores.html', user=user)

@app.route('/analise-potencial')
@login_required
def analise_potencial_page():
    user = auth_manager.get_current_user()
    return render_template('analise_potencial.html', user=user)

@app.route('/analise-territorial')
@login_required
def analise_territorial_page():
    user = auth_manager.get_current_user()
    return render_template('analise_territorial.html', user=user)

# API endpoints para Revendas
@app.route('/api/revendas', methods=['GET'])
@login_required
def get_revendas():
    """Recuperar todas as revendas ativas"""
    try:
        user = auth_manager.get_current_user()
        result = auth_manager.get_revendas()

        if result['success']:
            revendas_list = []
            for revenda in result['revendas']:
                # Convert dataclass to dict and add municipios_count
                if hasattr(revenda, '__dict__'):
                    revenda_dict = revenda.__dict__.copy()
                else:
                    revenda_dict = revenda.copy()

                # Convert datetime objects to strings to avoid serialization errors
                if revenda_dict.get('created_at'):
                    if hasattr(revenda_dict['created_at'], 'isoformat'):
                        revenda_dict['created_at'] = revenda_dict['created_at'].isoformat()

                if revenda_dict.get('updated_at'):
                    if hasattr(revenda_dict['updated_at'], 'isoformat'):
                        revenda_dict['updated_at'] = revenda_dict['updated_at'].isoformat()

                # Ensure municipios_codigos is properly handled
                municipios = revenda_dict.get('municipios_codigos', [])
                if not isinstance(municipios, list):
                    try:
                        import json
                        if isinstance(municipios, str):
                            municipios = json.loads(municipios)
                        else:
                            municipios = []
                    except:
                        municipios = []

                revenda_dict['municipios_codigos'] = municipios
                revenda_dict['municipios'] = municipios  # Add alias for compatibility
                revenda_dict['municipios_count'] = len(municipios)

                revendas_list.append(revenda_dict)

            return jsonify({
                'success': True,
                'revendas': revendas_list
            })
        else:
            return jsonify(result), 400

    except Exception as e:
        print(f"Erro na API de revendas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Erro ao recuperar revendas: {str(e)}'
        }), 500

@app.route('/api/revendas', methods=['POST'])
@login_required
def api_create_revenda():
    try:
        data = request.get_json()
        user = auth_manager.get_current_user()

        # Validação básica
        nome = data.get('nome', '').strip()
        cnpj = data.get('cnpj', '').strip()
        cnae = data.get('cnae', '').strip()
        cor = data.get('cor', '#4CAF50')
        municipios = data.get('municipios', [])

        if not nome:
            return jsonify({'success': False, 'error': 'Nome da revenda é obrigatório'})

        if not cnpj:
            return jsonify({'success': False, 'error': 'CNPJ é obrigatório'})

        if not cnae:
            return jsonify({'success': False, 'error': 'CNAE é obrigatório'})

        if notmunicipios or len(municipios) == 0:
            return jsonify({'success': False, 'error': 'Pelo menos um município deve ser selecionado'})

        # Garantir que municipios é uma lista
        if isinstance(municipios, str):
            municipios = [municipios]
        elif not isinstance(municipios, list):
            municipios = list(municipios) if municipios else []

        print(f"DEBUG: Criando revenda com municípios: {municipios}")

        # Dados da revenda
        revenda_data = {
            'nome': nome,
            'cnpj': cnpj,
            'cnae': cnae,
            'cor': cor,
            'municipios_codigos': municipios,
            'created_by': user['id']
        }

        result = auth_manager.create_revenda(revenda_data)

        if result['success']:
            return jsonify({'success': True, 'message': 'Revenda cadastrada com sucesso!'})
        else:
            return jsonify({'success': False, 'error': result['error']})

    except Exception as e:
        print(f"Erro ao criar revenda: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Erro interno do servidor'})

@app.route('/api/revendas/<int:revenda_id>', methods=['PUT'])
@login_required
def update_revenda(revenda_id):
    try:
        data = request.get_json()

        # Preparar dados para atualização no Supabase
        updates = {}
        if 'nome' in data:
            updates['nome'] = data['nome']
        if 'cnpj' in data:
            updates['cnpj'] = data['cnpj']
        if 'endereco' in data:
            updates['endereco'] = data['endereco']
        if 'cidade' in data:
            updates['cidade'] = data['cidade']
        if 'estado' in data:
            updates['estado'] = data['estado']
        if 'cep' in data:
            updates['cep'] = data['cep']
        if 'telefone' in data:
            updates['telefone'] = data['telefone']
        if 'email' in data:
            updates['email'] = data['email']
        if 'responsavel' in data:
            updates['responsavel'] = data['responsavel']

        result = auth_manager.update_revenda(revenda_id, updates)

        if result['success']:
            return jsonify({
                'success': True,
                'message': 'Revenda atualizada com sucesso!'
            })
        else:
            return jsonify({'success': False, 'error': result['error']}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/revendas/<int:revenda_id>', methods=['DELETE'])
@login_required
def delete_revenda(revenda_id):
    try:
        # Soft delete - marcar como inativo no Supabase
        result = auth_manager.delete_revenda(revenda_id)

        if result['success']:
            return jsonify({
                'success': True,
                'message': 'Revenda removida com sucesso!'
            })
        else:
            return jsonify({'success': False, 'error': result['error']}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/municipios/search')
@login_required
def search_municipios():
    try:
        query = request.args.get('q', '').lower()
        if len(query) < 2:
            return jsonify({'success': False, 'error': 'Query muito curta'}), 400

        # Buscar municípios nos dados de culturas para obter nomes reais
        municipios_found = []
        municipios_set = set()  # Para evitar duplicatas

        # Buscar nos dados de culturas
        for crop_name, crop_data in CROP_DATA.items():
            for municipality_code, municipality_data in crop_data.items():
                municipality_code_str = str(municipality_code)
                municipality_name = municipality_data.get('municipality_name', '').lower()
                state_code = municipality_data.get('state_code', 'XX')

                # Verificar se é um código de município válido e se contém a query
                if (len(municipality_code_str) == 7 and 
                    municipality_code_str.isdigit() and
                    municipality_code_str[0] in '12345' and
                    municipality_data.get('municipality_name') and
                    municipality_code not in municipios_set and
                    (query in municipality_name or query in state_code.lower())):

                    municipios_found.append({
                        'code': municipality_code,
                        'name': municipality_data.get('municipality_name'),
                        'state': state_code,
                        'full_name': f"{municipality_data.get('municipality_name')} ({state_code})"
                    })
                    municipios_set.add(municipality_code)

                    # Limitar a 50 resultados para performance
                    if len(municipios_found) >= 50:
                        break

            if len(municipios_found) >= 50:
                break

        # Ordenar por nome
        municipios_found.sort(key=lambda x: x['name'])

        return jsonify({
            'success': True,
            'municipios': municipios_found[:20]  # Limitar a 20 resultados
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/analise-potencial/<int:revenda_id>')
@login_required
def get_analise_potencial(revenda_id):
    try:
        # Ensure tables exist
        with app.app_context():
            db.create_all()

        # Tentar buscar no banco local primeiro
        revenda = Revenda.query.get(revenda_id)
        municipios_codes = []

        if revenda:
            print(f"Analyzing potential for revenda (local): {revenda.nome} (ID: {revenda_id})")
            municipios_codes = revenda.get_municipios_list()
        else:
            # Tentar buscar no Supabase
            try:
                result = auth_manager.supabase_manager.get_revenda_by_id(revenda_id)
                if result['success']:
                    revenda_data = result['data']
                    print(f"Analyzing potential for revenda (Supabase): {revenda_data.get('nome', 'Unknown')} (ID: {revenda_id})")
                    municipios_codes = []  # Implementar logic para municípios do Supabase
                else:
                    return jsonify({'success': False, 'error': 'Revenda não encontrada'}), 404
            except Exception as supabase_error:
                print(f"Supabase error: {supabase_error}")
                return jsonify({'success': False, 'error': 'Revenda não encontrada'}), 404

        # Obter dados de todas as fontes
        analysis_result = analyze_revenda_potential(municipios_codes)

        return jsonify({
            'success': True,
            'analysis': analysis_result
        })

    except Exception as e:
        print(f"Erro ao analisar potencial da revenda {revenda_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/analise-potencial-vendedor/<int:vendedor_id>')
@login_required
def get_analise_potencial_vendedor(vendedor_id):
    try:
        # Ensure tables exist
        with app.app_context():
            db.create_all()

        # Tentar buscar no banco local primeiro
        vendedor = Vendedor.query.get(vendedor_id)
        municipios_codes = []

        if vendedor:
            print(f"Analyzing potential for vendedor (local): {vendedor.nome} (ID: {vendedor_id})")
            municipios_codes = vendedor.get_municipios_list()
        else:
            # Tentar buscar no Supabase
            try:
                result = auth_manager.supabase_manager.get_vendedor_by_id(vendedor_id)
                if result['success']:
                    vendedor_data = result['data']
                    print(f"Analyzing potential for vendedor (Supabase): {vendedor_data.get('nome', 'Unknown')} (ID: {vendedor_id})")
                    municipios_codes = []  # Implementar logic para municípios do Supabase
                else:
                    return jsonify({'success': False, 'error': 'Vendedor não encontrado'}), 404
            except Exception as supabase_error:
                print(f"Supabase error: {supabase_error}")
                return jsonify({'success': False, 'error': 'Vendedor não encontrado'}), 404

        # Obter dados de todas as fontes
        analysis_result = analyze_revenda_potential(municipios_codes)

        return jsonify({
            'success': True,
            'analysis': analysis_result
        })

    except Exception as e:
        print(f"Erro ao analisar potencial do vendedor {vendedor_id}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

def analyze_revenda_potential(municipios_codes):
    """Analisa o potencial de uma revenda baseado em seus municípios"""

    analysis = {
        'potentialScore': 0,
        'totalValue': 0,
        'cropsDiversity': 0,
        'avgProductivity': 0,
        'fertilizersUsage': 0,
        'agrotoxicosUsage': 0,
        'technicalAssistance': 0,
        'recommendations': [],
        'dataBySource': {},
        'topCrops': [],
        'calculationMatrix': {}  # Nova matriz detalhada de cálculo
    }

    try:
        # Análise de Culturas
        crops_analysis = analyze_crops_data(municipios_codes)
        analysis['cropsDiversity'] = crops_analysis['diversity']
        analysis['avgProductivity'] = crops_analysis['avg_productivity']
        analysis['topCrops'] = crops_analysis['top_crops']
        analysis['dataBySource']['crops'] = crops_analysis['total_value']

        # Análise de Fertilizantes
        fertilizers_analysis = analyze_fertilizers_data(municipios_codes)
        analysis['fertilizersUsage'] = fertilizers_analysis['usage_percentage']
        analysis['dataBySource']['fertilizers'] = fertilizers_analysis['total_establishments']

        # Análise de Agrotóxicos
        agrotoxicos_analysis = analyze_agrotoxicos_data(municipios_codes)
        analysis['agrotoxicosUsage'] = agrotoxicos_analysis['usage_percentage']
        analysis['dataBySource']['agrotoxicos'] = agrotoxicos_analysis['total_establishments']

        # Análise de Consultoria Técnica
        consultoria_analysis = analyze_consultoria_data(municipios_codes)
        analysis['technicalAssistance'] = consultoria_analysis['coverage_percentage']
        analysis['dataBySource']['consultoria'] = consultoria_analysis['total_establishments']

        # Análise de Corretivos
        corretivos_analysis = analyze_corretivos_data(municipios_codes)
        analysis['dataBySource']['corretivos'] = corretivos_analysis['total_establishments']

        # Análise de Despesas
        despesas_analysis = analyze_despesas_data(municipios_codes)
        analysis['dataBySource']['despesas'] = despesas_analysis['total_value']

        # Análise de Escolaridade
        escolaridade_analysis = analyze_escolaridade_data(municipios_codes)
        analysis['dataBySource']['escolaridade'] = escolaridade_analysis['total_people']

        # Análise de Receitas
        receitas_analysis = analyze_receitas_data(municipios_codes)
        analysis['dataBySource']['receitas'] = receitas_analysis['total_value']

        # Calcular valor total
        analysis['totalValue'] = (
            analysis['dataBySource'].get('crops', 0) * 1000 +  # Estimativa de valor por hectare
            analysis['dataBySource'].get('despesas', 0) +
            analysis['dataBySource'].get('receitas', 0)
        )

        # NOVO ALGORITMO DE CÁLCULO MAIS PRECISO E DETALHADO
        calculation_matrix = {}

        # 1. Diversidade Agrícola (30% - máximo 30 pontos)
        diversity_raw = analysis['cropsDiversity']
        diversity_normalized = min(diversity_raw / 20.0, 1.0)  # Normalizar para 0-1 (20+ culturas = máximo)
        diversity_score = diversity_normalized * 30
        calculation_matrix['diversity'] = {
            'raw_value': diversity_raw,
            'normalized': diversity_normalized,
            'weight': 0.30,
            'max_points': 30,
            'score': diversity_score,
            'description': f'{diversity_raw} culturas diferentes encontradas'
        }

        # 2. Performance Financeira (25% - máximo 25 pontos)
        total_receita = analysis['dataBySource'].get('receitas', 0)
        total_despesa = analysis['dataBySource'].get('despesas', 0)
        saldo_financeiro = total_receita - total_despesa

        # Normalizar saldo financeiro (valores positivos são melhores)
        if total_receita > 0 or total_despesa > 0:
            financial_ratio = max(0, saldo_financeiro) / max(total_receita, 1)
            financial_normalized = min(financial_ratio, 1.0)
        else:
            financial_normalized = 0

        financial_score = financial_normalized * 25
        calculation_matrix['financial'] = {
            'raw_value': saldo_financeiro,
            'receita': total_receita,
            'despesa': total_despesa,
            'normalized': financial_normalized,
            'weight': 0.25,
            'max_points': 25,
            'score': financial_score,
            'description': f'Saldo: R$ {saldo_financeiro:,.2f} (Receita: R$ {total_receita:,.2f}, Despesa: R$ {total_despesa:,.2f})'
        }

        # 3. Abrangência Territorial (20% - máximo 20 pontos)
        num_municipios = len(municipios_codes)
        territorial_normalized = min(num_municipios / 50.0, 1.0)  # 50+ municípios = máximo
        territorial_score = territorial_normalized * 20
        calculation_matrix['territorial'] = {
            'raw_value': num_municipios,
            'normalized': territorial_normalized,
            'weight': 0.20,
            'max_points': 20,
            'score': territorial_score,
            'description': f'{num_municipios} municípios na área de atuação'
        }

        # 4. Atividade de Mercado (25% - máximo 25 pontos)
        # Subcritérios: Fertilizantes (40%), Consultoria (30%), Agrotóxicos (30%)
        fertilizers_subscore = (analysis['fertilizersUsage'] / 100.0) * 0.4
        consultoria_subscore = (analysis['technicalAssistance'] / 100.0) * 0.3
        agrotoxicos_subscore = (analysis['agrotoxicosUsage'] / 100.0) * 0.3

        market_activity_normalized = fertilizers_subscore + consultoria_subscore + agrotoxicos_subscore
        market_activity_score = market_activity_normalized * 25

        calculation_matrix['market_activity'] = {
            'raw_values': {
                'fertilizers': analysis['fertilizersUsage'],
                'consultoria': analysis['technicalAssistance'],
                'agrotoxicos': analysis['agrotoxicosUsage']
            },
            'subscores': {
                'fertilizers': fertilizers_subscore,
                'consultoria': consultoria_subscore,
                'agrotoxicos': agrotoxicos_subscore
            },
            'normalized': market_activity_normalized,
            'weight': 0.25,
            'max_points': 25,
            'score': market_activity_score,
            'description': f'Fertilizantes: {analysis["fertilizersUsage"]:.1f}%, Consultoria: {analysis["technicalAssistance"]:.1f}%, Agrotóxicos: {analysis["agrotoxicosUsage"]:.1f}%'
        }

        # Calcular pontuação final
        final_score = diversity_score + financial_score + territorial_score + market_activity_score

        # Aplicar ajustes baseados na produtividade média
        productivity_bonus = 0
        if analysis['avgProductivity'] > 0:
            # Bonus de até 5 pontos para alta produtividade
            productivity_normalized = min(analysis['avgProductivity'] / 5000.0, 1.0)  # 5000+ ha/município = máximo
            productivity_bonus = productivity_normalized * 5

        calculation_matrix['productivity_bonus'] = {
            'raw_value': analysis['avgProductivity'],
            'normalized': productivity_normalized if analysis['avgProductivity'] > 0 else 0,
            'bonus_points': productivity_bonus,
            'description': f'Produtividade média: {analysis["avgProductivity"]:.1f} ha/município'
        }

        final_score = min(final_score + productivity_bonus, 100)  # Máximo de 100 pontos

        # Resumo da matriz de cálculo
        calculation_matrix['summary'] = {
            'base_score': diversity_score + financial_score + territorial_score + market_activity_score,
            'productivity_bonus': productivity_bonus,
            'final_score': final_score,
            'breakdown': {
                'diversity': diversity_score,
                'financial': financial_score,
                'territorial': territorial_score,
                'market_activity': market_activity_score
            }
        }

        analysis['potentialScore'] = final_score
        analysis['calculationMatrix'] = calculation_matrix

        # Gerar recomendações baseadas na matriz de cálculo
        analysis['recommendations'] = generate_enhanced_recommendations(analysis, calculation_matrix)

    except Exception as e:
        print(f"Erro na análise de potencial: {str(e)}")
        import traceback
        traceback.print_exc()
        # Retornar análise padrão em caso de erro
        analysis['recommendations'] = ['Erro ao processar dados - análise limitada']
        analysis['calculationMatrix'] = {'error': str(e)}

    return analysis

def analyze_crops_data(municipios_codes):
    """Analisa dados de culturas para os municípios"""
    total_hectares = 0
    crops_found = set()
    top_crops = []

    try:
        for crop_name, crop_data in CROP_DATA.items():
            crop_total = 0
            for municipio_code in municipios_codes:
                municipio_str = str(municipio_code)
                if municipio_str in crop_data:
                    area = crop_data[municipio_str].get('harvested_area', 0)
                    total_hectares += area
                    crop_total += area
                    if area > 0:
                        crops_found.add(crop_name)

            if crop_total > 0:
                top_crops.append({'name': crop_name, 'value': crop_total})

        # Ordenar e pegar top 5 culturas
        top_crops.sort(key=lambda x: x['value'], reverse=True)
        top_crops = top_crops[:5]

    except Exception as e:
        print(f"Erro na análise de culturas: {str(e)}")

    return {
        'diversity': len(crops_found),
        'avg_productivity': total_hectares / max(len(municipios_codes), 1),
        'total_value': total_hectares,
        'top_crops': top_crops
    }

def analyze_fertilizers_data(municipios_codes):
    """Analisa dados de fertilizantes para os municípios"""
    total_establishments = 0
    municipalities_with_data = 0

    try:
        # Excluir categorias que são totalizações
        excluded_categories = ['Total Estabelecimentos', 'Total estabelecimentos', 'TOTAL']

        for category_name, category_data in FERTILIZER_DATA.items():
            # Pular categorias de totalização
            if category_name in excluded_categories:
                continue

            for municipio_code in municipios_codes:
                municipio_str = str(municipio_code)
                if municipio_str in category_data:
                    value = category_data[municipio_str].get('value', 0)
                    total_establishments += value
                    if value > 0:
                        municipalities_with_data += 1
                        break  # Conta o município apenas uma vez
    except Exception as e:
        print(f"Erro na análise de fertilizantes: {str(e)}")

    usage_percentage = (municipalities_with_data / max(len(municipios_codes), 1)) * 100

    return {
        'total_establishments': total_establishments,
        'usage_percentage': min(usage_percentage, 100)
    }

def analyze_agrotoxicos_data(municipios_codes):
    """Analisa dados de agrotóxicos para os municípios"""
    total_establishments = 0
    municipalities_with_data = 0

    try:
        for category_name, category_data in AGROTOXICO_DATA.items():
            for municipio_code in municipios_codes:
                municipio_str = str(municipio_code)
                if municipio_str in category_data:
                    value = category_data[municipio_str].get('value', 0)
                    total_establishments += value
                    if value > 0:
                        municipalities_with_data += 1
                        break
    except Exception as e:
        print(f"Erro na análise de agrotóxicos: {str(e)}")

    usage_percentage = (municipalities_with_data / max(len(municipios_codes), 1)) * 100

    return {
        'total_establishments': total_establishments,
        'usage_percentage': min(usage_percentage, 100)
    }

def analyze_consultoria_data(municipios_codes):
    """Analisa dados de consultoria técnica para os municípios"""
    total_establishments = 0
    municipalities_with_data = 0

    try:
        for category_name, category_data in CONSULTORIA_DATA.items():
            for municipio_code in municipios_codes:
                municipio_str = str(municipio_code)
                if municipio_str in category_data:
                    value = category_data[municipio_str].get('value', 0)
                    total_establishments += value
                    if value > 0:
                        municipalities_with_data += 1
                        break
    except Exception as e:
        print(f"Erro na análise de consultoria: {str(e)}")

    coverage_percentage = (municipalities_with_data / max(len(municipios_codes), 1)) * 100

    return {
        'total_establishments': total_establishments,
        'coverage_percentage': min(coverage_percentage, 100)
    }

def analyze_corretivos_data(municipios_codes):
    """Analisa dados de corretivos para os municípios"""
    total_establishments = 0

    try:
        for category_name, category_data in CORRETIVOS_DATA.items():
            for municipio_code in municipios_codes:
                municipio_str = str(municipio_code)
                if municipio_str in category_data:
                    value = category_data[municipio_str].get('value', 0)
                    total_establishments += value
    except Exception as e:
        print(f"Erro na análise de corretivos: {str(e)}")

    return {
        'total_establishments': total_establishments
    }

def analyze_despesas_data(municipios_codes):
    """Analisa dados de despesas para os municípios"""
    total_value = 0

    try:
        for category_name, category_data in DESPESA_DATA.items():
            for municipio_code in municipios_codes:
                municipio_str = str(municipio_code)
                if municipio_str in category_data:
                    value = category_data[municipio_str].get('value', 0)
                    total_value += value
    except Exception as e:
        print(f"Erro na análise de despesas: {str(e)}")

    return {
        'total_value': total_value
    }

def analyze_escolaridade_data(municipios_codes):
    """Analisa dados de escolaridade para os municípios"""
    total_people = 0

    try:
        for category_name, category_data in ESCOLARIDADE_DATA.items():
            for municipio_code in municipios_codes:
                municipio_str = str(municipio_code)
                if municipio_str in category_data:
                    value = category_data[municipio_str].get('value', 0)
                    total_people += value
    except Exception as e:
        print(f"Erro na análise de escolaridade: {str(e)}")

    return {
        'total_people': total_people
    }

def analyze_receitas_data(municipios_codes):
    """Analisa dados de receitas para os municípios"""
    total_value = 0

    try:
        for category_name, category_data in RECEITA_DATA.items():
            for municipio_code in municipios_codes:
                municipio_str = str(municipio_code)
                if municipio_str in category_data:
                    value = category_data[municipio_str].get('value', 0)
                    total_value += value
    except Exception as e:
        print(f"Erro na análise de receitas: {str(e)}")

    return {
        'total_value': total_value
    }

def generate_enhanced_recommendations(analysis, calculation_matrix):
    """Gera recomendações baseadas na análise detalhada"""
    recommendations = []

    try:
        # Análise baseada na matriz de cálculo
        matrix = calculation_matrix

        # Recomendações baseadas na diversidade
        if 'diversity' in matrix:
            diversity_score = matrix['diversity']['score']
            if diversity_score < 10:
                recommendations.append(f"⚠️ Baixa diversidade ({matrix['diversity']['raw_value']} culturas) - diversifique para reduzir riscos e aumentar {30-diversity_score:.1f} pontos")
            elif diversity_score > 25:
                recommendations.append(f"✅ Excelente diversidade ({matrix['diversity']['raw_value']} culturas) - mantenha essa estratégia")

        # Recomendações baseadas na performance financeira
        if 'financial' in matrix:
            financial_score = matrix['financial']['score']
            saldo = matrix['financial']['raw_value']
            if financial_score < 10:
                if saldo < 0:
                    recommendations.append(f"⚠️ Saldo financeiro negativo (R$ {saldo:,.2f}) - foque em eficiência e redução de custos")
                else:
                    recommendations.append(f"⚠️ Baixa performance financeira - oportunidade de crescimento de até {25-financial_score:.1f} pontos")
            elif financial_score > 20:
                recommendations.append(f"✅ Boa saúde financeira (saldo: R$ {saldo:,.2f}) - mantenha o equilíbrio")

        # Recomendações baseadas na abrangência territorial
        if 'territorial' in matrix:
            territorial_score = matrix['territorial']['score']
            num_municipios = matrix['territorial']['raw_value']
            if territorial_score < 8:
                recommendations.append(f"📍 Abrangência limitada ({num_municipios} municípios) - considere expansão territorial para ganhar {20-territorial_score:.1f} pontos")
            elif territorial_score > 15:
                recommendations.append(f"✅ Boa cobertura territorial ({num_municipios} municípios) - aproveite a escala")

        # Recomendações baseadas na atividade de mercado
        if 'market_activity' in matrix:
            market_score = matrix['market_activity']['score']
            raw_values = matrix['market_activity']['raw_values']

            if market_score < 10:
                weak_areas = []
                if raw_values['fertilizers'] < 30:
                    weak_areas.append("fertilizantes")
                if raw_values['consultoria'] < 25:
                    weak_areas.append("consultoria técnica")
                if raw_values['agrotoxicos'] < 20:
                    weak_areas.append("agrotóxicos")

                if weak_areas:
                    recommendations.append(f"📈 Baixa atividade de mercado em: {', '.join(weak_areas)} - oportunidade de {25-market_score:.1f} pontos")
            elif market_score > 20:
                recommendations.append("✅ Alto nível de atividade de mercado - mercado maduro e ativo")

        # Recomendações baseadas no bonus de produtividade
        if 'productivity_bonus' in matrix:
            bonus = matrix['productivity_bonus']['bonus_points']
            productivity = matrix['productivity_bonus']['raw_value']
            if bonus < 2:
                recommendations.append(f"⚡ Produtividade média baixa ({productivity:.1f} ha/município) - potencial de melhoria")
            elif bonus > 4:
                recommendations.append(f"✅ Alta produtividade média ({productivity:.1f} ha/município) - região muito produtiva")

        # Recomendação geral baseada na pontuação final
        final_score = analysis['potentialScore']
        if final_score >= 80:
            recommendations.append("🎯 EXCELENTE: Região prioritária para investimento e expansão")
        elif final_score >= 60:
            recommendations.append("👍 BOM: Região promissora com oportunidades sólidas")
        elif final_score >= 40:
            recommendations.append("⚖️ MÉDIO: Região com potencial, requer estratégia focada")
        else:
            recommendations.append("⚠️ BAIXO: Região desafiadora, considere estratégias alternativas")

        # Identificar maior oportunidade de melhoria
        if 'summary' in matrix:
            breakdown = matrix['summary']['breakdown']
            lowest_component = min(breakdown.items(), key=lambda x: x[1])
            component_names = {
                'diversity': 'diversidade de culturas',
                'financial': 'performance financeira',
                'territorial': 'abrangência territorial',
                'market_activity': 'atividade de mercado'
            }
            recommendations.append(f"🔍 PRIORIDADE: Melhorar {component_names.get(lowest_component[0], lowest_component[0])} ({lowest_component[1]:.1f} pts)")

        if not recommendations:
            recommendations.append("Continue monitorando os dados para identificar oportunidades")

    except Exception as e:
        print(f"Erro ao gerar recomendações: {str(e)}")
        recommendations = ["Erro ao gerar recomendações específicas"]

    return recommendations

@app.route('/api/revendas/data/<int:revenda_id>')
@login_required
def get_revenda_data(revenda_id):
    """Obter dados de território da revenda para visualização no mapa"""
    try:
        print(f"Loading territory data for revenda (Supabase): {revenda_id}")

        # Buscar revenda no Supabase
        result = auth_manager.get_revenda_by_id(revenda_id)

        if not result['success']:
            print(f"Revenda not found: {result['error']}")
            return jsonify({'success': False, 'error': 'Revenda não encontrada'})

        revenda_data = result['data']
        revenda_name = revenda_data.get('nome', f'Revenda {revenda_id}')
        municipios_codigos = revenda_data.get('municipios_codigos', [])

        print(f"Loading territory data for revenda (Supabase): {revenda_name} (ID: {revenda_id})")
        print(f"Raw municipios_codigos from database: {municipios_codigos}")
        print(f"Type of municipios_codigos: {type(municipios_codigos)}")

        # Garantir que municipios_codigos seja uma lista válida
        if not municipios_codigos:
            print("No municipalities found for this revenda - empty municipios_codigos")
            return jsonify({'success': False, 'error': 'Nenhum município encontrado para esta revenda'})

        if not isinstance(municipios_codigos, list):
            print(f"municipios_codigos is not a list, attempting to convert: {municipios_codigos}")
            try:
                import json
                if isinstance(municipios_codigos, str):
                    municipios_codigos = json.loads(municipios_codigos)
                else:
                    municipios_codigos = list(municipios_codigos) if municipios_codigos else []
            except Exception as conv_error:
                print(f"Error converting municipios_codigos: {conv_error}")
                return jsonify({'success': False, 'error': 'Formato inválido dos códigos de municípios'})

        print(f"Final municipios_codigos after processing: {municipios_codigos}")

        if not municipios_codigos or len(municipios_codigos) == 0:
            print("No municipalities found for this revenda after processing")
            return jsonify({'success': False, 'error': 'Nenhum município encontrado para esta revenda'})

        # Preparar dados dos municípios
        territory_data = {}

        # Para cada município do território da revenda
        for municipio_code in municipios_codigos:
            municipio_code_str = str(municipio_code)

            # Buscar nome do município nos dados de culturas
            municipality_name = f"Município {municipio_code}" # Default name
            state_code = "XX" # Default state code

            # Procurar o nome real do município nos dados de culturas
            for crop_name, crop_data in CROP_DATA.items():
                if municipio_code_str in crop_data:
                    municipality_data = crop_data[municipio_code_str]
                    municipality_name = municipality_data.get('municipality_name', municipality_name)
                    state_code = municipality_data.get('state_code', state_code)
                    break

            territory_data[municipio_code_str] = {
                'municipality_name': municipality_name,
                'state_code': state_code,
                'value': 1,  # Valor fixo para território
                'unit': 'território'
            }

        print(f"Territory data prepared for {len(territory_data)} municipalities")

        return jsonify({
            'success': True,
            'data': territory_data,
            'revenda_name': revenda_name
        })

    except Exception as e:
        print(f"Error loading revenda territory data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Erro ao carregar dados do território: {str(e)}'
        })

# Template e Upload para Revendas
@app.route('/api/revendas/template')
def download_revendas_template():
    """Download template Excel para cadastro de revendas"""
    try:
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cadastro de Revendas"

        # Cabeçalhos com formatação
        headers = [
            'Nome da Revenda', 'CNPJ', 'CNAE Principal', 'Cor (Hex)',
            'Código IBGE Município', 'Nome do Município', 'UF'
        ]

        # Aplicar formatação nos cabeçalhos
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Instruções na segunda linha
        instructions = [
            'INSTRUÇÕES: Preencha os dados da revenda na primeira linha válida.',
            'Para adicionar múltiplos municípios, copie a linha da revenda alterando apenas o município.',
            'Use códigos IBGE válidos - baixe a planilha de referência.',
            'Cores em formato hexadecimal (#RRGGBB).',
            'CNAE deve estar no formato 0000-0/00.',
            '', ''
        ]

        for col, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=2, column=col, value=instruction)
            cell.font = Font(italic=True, color="666666")

        # Exemplo de linha com dados reais
        example_data = [
            'Agro Fertilizantes Ltda', '12.345.678/0001-90', '4681-8/01', '#4CAF50',
            '3550308', 'São Paulo', 'SP'
        ]

        for col, data in enumerate(example_data, 1):
            ws.cell(row=3, column=col, value=data)

        # Segundo exemplo para mostrar múltiplos municípios
        example_data2 = [
            'Agro Fertilizantes Ltda', '12.345.678/0001-90', '4681-8/01', '#4CAF50',
            '3304557', 'Rio de Janeiro', 'RJ'
        ]

        for col, data in enumerate(example_data2, 1):
            ws.cell(row=4, column=col, value=data)

        # Ajustar largura das colunas
        column_widths = [30, 20, 15, 12, 20, 25, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Adicionar bordas
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # Salvar em memória
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='template_cadastro_revendas.xlsx'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/municipios/referencia')
def download_municipios_referencia():
    """Download planilha de referência com todos os códigos IBGE de municípios"""
    try:
        # Carregar dados dos municípios do arquivo estático
        crop_data_path = os.path.join(os.path.dirname(__file__), 'data', 'crop_data_static.json')

        municipios_data = []

        if os.path.exists(crop_data_path):
            with open(crop_data_path, 'r', encoding='utf-8') as f:
                crop_data = json.load(f)

            # Extrair todos os municípios únicos de todas as culturas
            all_municipalities = {}

            for crop_name, crop_municipalities in crop_data.items():
                for municipality_code, municipality_data in crop_municipalities.items():
                    municipality_code_str = str(municipality_code)
                    municipality_name = municipality_data.get('municipality_name', '').lower()

                    # Verificar se é um código de município válido
                    if (len(municipality_code_str) == 7 and 
                        municipality_code_str.isdigit() and
                        municipality_code_str[0] in '12345' and
                        municipality_data.get('municipality_name') and
                        not any(keyword in municipality_name for keyword in [
                            'região', 'mesorregião', 'microrregião', 'nordeste', 'norte', 'sul', 
                            'centro', 'oeste', 'leste', 'sudeste', 'noroeste', 'sudoeste',
                            'alto ', 'baixo ', 'médio ', '-grossense', 'parecis', 'araguaia',
                            'pantanal', 'cerrado', 'amazônia', 'caatinga', 'mata atlântica'
                        ]) and
                        municipality_name not in [
                            'alto teles pires', 'sudeste mato-grossense', 'parecis', 'barreiras',
                            'dourados', 'norte mato-grossense', 'portal da amazônia'
                        ]):

                        all_municipalities[municipality_code] = {
                            'codigo_municipio': municipality_code,
                            'nome_municipio': municipality_data.get('municipality_name', ''),
                            'uf': municipality_data.get('state_code', '')
                        }

            # Converter para lista e ordenar
            municipios_data = list(all_municipalities.values())
            municipios_data.sort(key=lambda x: (x['uf'], x['nome_municipio']))

        # Se não conseguir carregar os dados, usar alguns exemplos
        if not municipios_data:
            municipios_data = [
                {'codigo_municipio': '3550308', 'nome_municipio': 'São Paulo', 'uf': 'SP'},
                {'codigo_municipio': '3304557', 'nome_municipio': 'Rio de Janeiro', 'uf': 'RJ'},
                {'codigo_municipio': '3106200', 'nome_municipio': 'Belo Horizonte', 'uf': 'MG'},
                {'codigo_municipio': '2304400', 'nome_municipio': 'Fortaleza', 'uf': 'CE'},
                {'codigo_municipio': '4106902', 'nome_municipio': 'Curitiba', 'uf': 'PR'},
            ]

        # Criar workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Municípios IBGE"

        # Cabeçalhos
        headers = ['Código IBGE', 'Nome do Município', 'UF']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Dados dos municípios
        for row, municipio in enumerate(municipios_data, 2):
            ws.cell(row=row, column=1, value=municipio['codigo_municipio'])
            ws.cell(row=row, column=2, value=municipio['nome_municipio'])
            ws.cell(row=row, column=3, value=municipio['uf'])

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 8

        # Adicionar filtros
        ws.auto_filter.ref = f"A1:C{len(municipios_data) + 1}"

        # Salvar em memória
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='municipios_ibge_referencia.xlsx'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Template para Vendedores
@app.route('/api/vendedores/template')
def download_vendedores_template():
    """Download template Excel para cadastro de vendedores"""
    try:
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cadastro de Vendedores"

        # Cabeçalhos com formatação
        headers = [
            'Nome Completo', 'E-mail', 'Telefone', 'CPF', 'Cor (Hex)',
            'Código IBGE Município', 'Nome do Município', 'UF'
        ]

        # Aplicar formatação nos cabeçalhos
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Instruções na segunda linha
        instructions = [
            'INSTRUÇÕES: Preencha os dados do vendedor na primeira linha válida.',
            'Para adicionar múltiplos municípios, copie a linha do vendedor alterando apenas o município.',
            'Use códigos IBGE válidos - baixe a planilha de referência.',
            'Cores em formato hexadecimal (#RRGGBB).',
            'Mantenha este formato para upload correto.',
            '', '', ''
        ]

        for col, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=2, column=col, value=instruction)
            cell.font = Font(italic=True, color="666666")

        # Exemplo de linha com dados reais
        example_data = [
            'João Silva Santos', 'joao.silva@email.com', '(11) 99999-8888', '123.456.789-00', '#2196F3',
            '3550308', 'São Paulo', 'SP'
        ]

        for col, data in enumerate(example_data, 1):
            ws.cell(row=3, column=col, value=data)

        # Segundo exemplo para mostrar múltiplos municípios
        example_data2 = [
            'João Silva Santos', 'joao.silva@email.com', '(11) 99999-8888', '123.456.789-00', '#2196F3',
            '3304557', 'Rio de Janeiro', 'RJ'
        ]

        for col, data in enumerate(example_data2, 1):
            ws.cell(row=4, column=col, value=data)

        # Ajustar largura das colunas
        column_widths = [25, 30, 18, 18, 12, 20, 25, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Adicionar bordas
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # Salvar em memória
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='template_cadastro_vendedores.xlsx'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/revendas/upload', methods=['POST'])
@login_required
def upload_revendas_file():
    """Process uploaded Excel file for revendas"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'}), 400

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400

        # Ler arquivo Excel
        df = pd.read_excel(file)

        # Validar colunas obrigatórias
        required_columns = ['Nome da Revenda', 'CNPJ', 'CNAE Principal', 'Código IBGE Município']
        missing_columns = []
        for col in required_columns:
            if col not in df.columns:
                missing_columns.append(col)

        if missing_columns:
            return jsonify({'success': False, 'error': f'Colunas obrigatórias não encontradas: {", ".join(missing_columns)}'}), 400

        print(f"DEBUG: Colunas encontradas no Excel: {list(df.columns)}")
        print(f"DEBUG: Primeiras linhas do arquivo:")
        print(df.head(3).to_string())

        # Filtrar linhas válidas (pular instruções)
        df_valid = df[df['Nome da Revenda'].notna() & 
                     df['CNPJ'].notna() & 
                     df['CNAE Principal'].notna() & 
                     df['Código IBGE Município'].notna() &
                     ~df['Nome da Revenda'].astype(str).str.contains('INSTRUÇÕES|Preencha|Os dados|Para adicionar', case=False, na=False)]

        print(f"DEBUG: Linhas válidas encontradas: {len(df_valid)}")

        if len(df_valid) == 0:
            return jsonify({'success': False, 'error': 'Nenhuma linha válida encontrada no arquivo'}), 400

        # Extrair dados da revenda (da primeira linha válida)
        first_row = df_valid.iloc[0]
        revenda_data = {
            'nome': str(first_row['Nome da Revenda']).strip(),
            'cnpj': str(first_row['CNPJ']).strip(),
            'cnae': str(first_row['CNAE Principal']).strip(),
            'cor': str(first_row.get('Cor (Hex)', '#4CAF50')).strip()
        }

        # Validar CNPJ único no Supabase
        try:
            existing_check = auth_manager.supabase_manager.supabase.table('revendas').select('*').eq('cnpj', revenda_data['cnpj']).eq('active', True).execute()
            if existing_check.data:
                return jsonify({'success': False, 'error': f'CNPJ {revenda_data["cnpj"]} já está cadastrado'}), 400
        except Exception as check_error:
            print(f"Erro ao verificar CNPJ existente: {check_error}")

        # Extrair municípios
        municipios = []
        for _, row in df_valid.iterrows():
            municipio_code = str(row['Código IBGE Município']).strip()
            municipio_name = str(row.get('Nome do Município', 'N/A')).strip()
            uf = str(row.get('UF', 'XX')).strip()

            if municipio_code and municipio_code != 'nan':
                municipios.append({
                    'code': municipio_code,
                    'name': municipio_name,
                    'uf': uf
                })

        return jsonify({
            'success': True,
            'data': {
                'revenda': revenda_data,
                'municipios': municipios,
                'total_municipios': len(municipios)
            }
        })

    except Exception as e:
        print(f"Error processing upload: {str(e)}")
        return jsonify({'success': False, 'error': f'Erro ao processar arquivo: {str(e)}'}), 500

@app.route('/api/vendedores/upload', methods=['POST'])
@login_required
def upload_vendedores_file():
    """Process uploaded Excel file for vendedores"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Nenhum arquivo enviado'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Nenhum arquivo selecionado'}), 400

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400

        # Ler arquivo Excel
        df = pd.read_excel(file)

        # Validar colunas obrigatórias
        required_columns = ['Nome Completo', 'E-mail', 'Telefone', 'CPF', 'Código IBGE Município']
        for col in required_columns:
            if col not in df.columns:
                return jsonify({'success': False, 'error': f'Coluna obrigatória não encontrada: {col}'}), 400

        # Filtrar linhas válidas (pular instruções)
        df_valid = df[df['Nome Completo'].notna() & 
                     df['E-mail'].notna() & 
                     df['Telefone'].notna() & 
                     df['CPF'].notna() &
                     df['Código IBGE Município'].notna() &
                     ~df['Nome Completo'].str.contains('INSTRUÇÕES|Preencha|Os dados|Para adicionar', na=False)]

        if len(df_valid) == 0:
            return jsonify({'success': False, 'error': 'Nenhuma linha válida encontrada no arquivo'}), 400

        # Extrair dados do vendedor (da primeira linha válida)
        first_row = df_valid.iloc[0]
        vendedor_data = {
            'nome': str(first_row['Nome Completo']).strip(),
            'email': str(first_row['E-mail']).strip(),
            'telefone': str(first_row['Telefone']).strip(),
            'cpf': str(first_row['CPF']).strip(),
            'cor': str(first_row.get('Cor (Hex)', '#2196F3')).strip()
        }

        # Validar email e CPF únicos
        existing_email = Vendedor.query.filter_by(email=vendedor_data['email']).first()
        if existing_email:
            return jsonify({'success': False, 'error': f'E-mail {vendedor_data["email"]} já está cadastrado'}), 400

        existing_cpf = Vendedor.query.filter_by(cpf=vendedor_data['cpf']).first()
        if existing_cpf:
            return jsonify({'success': False, 'error': f'CPF {vendedor_data["cpf"]} já está cadastrado'}), 400

        # Extrair municípios
        municipios = []
        for _, row in df_valid.iterrows():
            municipio_code = str(row['Código IBGE Município']).strip()
            municipio_name = str(row.get('Nome do Município', 'N/A')).strip()
            uf = str(row.get('UF', 'XX')).strip()

            if municipio_code and municipio_code != 'nan':
                municipios.append({
                    'code': municipio_code,
                    'name': municipio_name,
                    'uf': uf
                })

        return jsonify({
            'success': True,
            'data': {
                'vendedor': vendedor_data,
                'municipios': municipios,
                'total_municipios': len(municipios)
            }
        })

    except Exception as e:
        print(f"Error processing upload: {str(e)}")
        return jsonify({'success': False, 'error': f'Erro ao processar arquivo: {str(e)}'}), 500

# API endpoints para Vendedores
@app.route('/api/vendedores', methods=['GET'])
@login_required
def get_vendedores():
    """Recuperar todos os vendedores ativos"""
    try:
        user = auth_manager.get_current_user()
        result = auth_manager.get_vendedores()

        if result['success']:
            vendedores_list = []
            for vendedor in result['vendedores']:
                # Convert dataclass to dict and add municipios_count
                if hasattr(vendedor, '__dict__'):
                    vendedor_dict = vendedor.__dict__.copy()
                else:
                    vendedor_dict = vendedor.copy()

                # Convert datetime objects to strings to avoid serialization errors
                if vendedor_dict.get('created_at'):
                    if hasattr(vendedor_dict['created_at'], 'isoformat'):
                        vendedor_dict['created_at'] = vendedor_dict['created_at'].isoformat()

                if vendedor_dict.get('updated_at'):
                    if hasattr(vendedor_dict['updated_at'], 'isoformat'):
                        vendedor_dict['updated_at'] = vendedor_dict['updated_at'].isoformat()

                # Ensure municipios_codigos is properly handled
                municipios = vendedor_dict.get('municipios_codigos', [])
                if not isinstance(municipios, list):
                    try:
                        import json
                        if isinstance(municipios, str):
                            municipios = json.loads(municipios)
                        else:
                            municipios = []
                    except:
                        municipios = []

                vendedor_dict['municipios_codigos'] = municipios
                vendedor_dict['municipios'] = municipios  # Add alias for compatibility
                vendedor_dict['municipios_count'] = len(municipios)

                vendedores_list.append(vendedor_dict)

            return jsonify({
                'success': True,
                'vendedores': vendedores_list
            })
        else:
            return jsonify(result), 400

    except Exception as e:
        print(f"Erro na API de vendedores: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Erro ao recuperar vendedores: {str(e)}'
        }), 500

@app.route('/api/vendedores', methods=['POST'])
@login_required
def api_create_vendedor():
    try:
        data = request.get_json()
        user = auth_manager.get_current_user()

        # Validação básica
        nome = data.get('nome', '').strip()
        email = data.get('email', '').strip()
        telefone = data.get('telefone', '').strip()
        cpf = data.get('cpf', '').strip()
        cor = data.get('cor', '#2196F3')
        municipios = data.get('municipios_codigos', [])

        if not nome:
            return jsonify({'success': False, 'error': 'Nome do vendedor é obrigatório'})

        if not email:
            return jsonify({'success': False, 'error': 'E-mail é obrigatório'})

        if not telefone:
            return jsonify({'success': False, 'error': 'Telefone é obrigatório'})

        if not cpf:
            return jsonify({'success': False, 'error': 'CPF é obrigatório'})

        if not municipios or len(municipios) == 0:
            return jsonify({'success': False, 'error': 'Pelo menos um município deve ser selecionado'})

        # Garantir que municipios é uma lista
        if isinstance(municipios, str):
            municipios = [municipios]
        elif not isinstance(municipios, list):
            municipios = list(municipios) if municipios else []

        print(f"DEBUG: Criando vendedor com municípios: {municipios}")

        # Dados do vendedor
        vendedor_data = {
            'nome': nome,
            'email': email,
            'telefone': telefone,
            'cpf': cpf,
            'cor': cor,
            'municipios_codigos': municipios,
            'created_by': user['id']
        }

        result = auth_manager.create_vendedor(vendedor_data)

        if result['success']:
            return jsonify({'success': True, 'message': 'Vendedor cadastrado com sucesso!'})
        else:
            return jsonify({'success': False, 'error': result['error']})

    except Exception as e:
        print(f"Erro ao criar vendedor: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': 'Erro interno do servidor'})

@app.route('/api/vendedores/<int:vendedor_id>', methods=['PUT'])
@login_required
def update_vendedor(vendedor_id):
    try:
        data = request.get_json()

        # Preparar dados para atualização no Supabase
        updates = {}
        if 'nome' in data:
            updates['nome'] = data['nome']
        if 'email' in data:
            updates['email'] = data['email']
        if 'telefone' in data:
            updates['telefone'] = data['telefone']
        if 'cor' in data:
            updates['cor'] = data['cor']
        if 'municipios_codigos' in data:
            updates['municipios_codigos'] = data['municipios_codigos']

        result = auth_manager.update_vendedor(vendedor_id, updates)

        if result['success']:
            return jsonify({
                'success': True,
                'message': 'Vendedor atualizado com sucesso!'
            })
        else:
            return jsonify({'success': False, 'error': result['error']}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/vendedores/<int:vendedor_id>', methods=['DELETE'])
@login_required
def delete_vendedor(vendedor_id):
    try:
        # Soft delete - marcar como inativo no Supabase
        result = auth_manager.delete_vendedor(vendedor_id)

        if result['success']:
            return jsonify({
                'success': True,
                'message': 'Vendedor removido com sucesso!'
            })
        else:
            return jsonify({'success': False, 'error': result['error']}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/revendas/create-from-upload', methods=['POST'])
@login_required
def create_revenda_from_upload():
    """Cria revenda a partir dos dados do upload Excel"""
    try:
        data = request.get_json()
        user = auth_manager.get_current_user()

        if not data:
            return jsonify({'success': False, 'error': 'Nenhum dado recebido'}), 400

        revenda_data = data.get('revenda', {})
        municipios_data = data.get('municipios', [])

        print(f"DEBUG routes: Criando revenda do upload com dados: {revenda_data}")
        print(f"DEBUG routes: Municípios do upload: {len(municipios_data)} itens")

        if not revenda_data.get('nome'):
            return jsonify({'success': False, 'error': 'Nome da revenda é obrigatório'}), 400

        if not municipios_data or len(municipios_data) == 0:
            return jsonify({'success': False, 'error': 'Nenhum município encontrado'}), 400

        # Extrair apenas os códigos dos municípios
        municipios_codigos = [m.get('code') for m in municipios_data if m.get('code')]

        print(f"DEBUG routes: Códigos de municípios extraídos: {municipios_codigos}")

        # Preparar dados para criação
        revenda_create_data = {
            'nome': revenda_data['nome'],
            'cnpj': revenda_data['cnpj'],
            'cnae': revenda_data['cnae'],
            'cor': revenda_data['cor'],
            'municipios_codigos': municipios_codigos,
            'created_by': user['id']
        }

        print(f"DEBUG routes: Dados finais para criação: {revenda_create_data}")

        # Criar revenda via Supabase
        result = auth_manager.create_revenda(revenda_create_data)

        if result['success']:
            return jsonify({
                'success': True,
                'message': f'Revenda "{revenda_data["nome"]}" cadastrada com sucesso! {len(municipios_codigos)} municípios adicionados.',
                'revenda_id': result['data'][0]['id'] if result['data'] else None
            })
        else:
            return jsonify({'success': False, 'error': result['error']})

    except Exception as e:
        print(f"Erro ao criar revenda do upload: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/vendedores/create-from-upload', methods=['POST'])
@login_required
def create_vendedor_from_upload():
    """Cria vendedor a partir dos dados do upload Excel"""
    try:
        data = request.get_json()
        user = auth_manager.get_current_user()

        if not data:
            return jsonify({'success': False, 'error': 'Nenhum dado recebido'}), 400

        vendedor_data = data.get('vendedor', {})
        municipios_data = data.get('municipios', [])

        print(f"DEBUG routes: Criando vendedor do upload com dados: {vendedor_data}")
        print(f"DEBUG routes: Municípios do upload: {len(municipios_data)} itens")

        if not vendedor_data.get('nome'):
            return jsonify({'success': False, 'error': 'Nome do vendedor é obrigatório'}), 400

        if not municipios_data or len(municipios_data) == 0:
            return jsonify({'success': False, 'error': 'Nenhum município encontrado'}), 400

        # Extrair apenas os códigos dos municípios
        municipios_codigos = [m.get('code') for m in municipios_data if m.get('code')]

        print(f"DEBUG routes: Códigos de municípios extraídos: {municipios_codigos}")

        # Preparar dados para criação
        vendedor_create_data = {
            'nome': vendedor_data['nome'],
            'email': vendedor_data.get('email', ''),
            'telefone': vendedor_data.get('telefone', ''),
            'cpf': vendedor_data.get('cpf', ''),
            'cor': vendedor_data.get('cor', '#2196F3'),
            'municipios_codigos': municipios_codigos,
            'created_by': user['id']
        }

        print(f"DEBUG routes: Dados finais para criação: {vendedor_create_data}")

        # Criar vendedor via Supabase
        result = auth_manager.create_vendedor(vendedor_create_data)

        if result['success']:
            return jsonify({
                'success': True,
                'message': f'Vendedor "{vendedor_data["nome"]}" cadastrado com sucesso! {len(municipios_codigos)} municípios adicionados.',
                'vendedor_id': result['data'][0]['id'] if result['data'] else None
            })
        else:
            return jsonify({'success': False, 'error': result['error']})

    except Exception as e:
        print(f"Erro ao criar vendedor do upload: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# Import models
from models import Revenda, Vendedor

# Create tables
with app.app_context():
    db.create_all()
    print("Tabelas criadas com sucesso!")

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)